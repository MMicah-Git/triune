"""Slice 1 validation — can we recover per-instance neck sizes from the plan
text layer by attaching the nearest size callout to each detection (using the
detection's already-known tag)? Scored against the completed takeoff's per-tag
size sets. Standalone, no pipeline changes."""
import fitz, re, json, glob, math, sys
from collections import defaultdict, Counter

DPI_TO_PT = 72.0 / 200
PROX_PT = 150               # pt; callouts sit ~20-120pt from the symbol center

# Size callout patterns (this engineer: 8"Ø round, 12"X10" rect with quotes)
ROUND = re.compile(r'^(\d+(?:\.\d+)?)\s*["″′]\s*[øØ⌀]?$')
RECT  = re.compile(r'^(\d+)\s*["″]?\s*[xX]\s*(\d+)\s*["″]?$')


def parse_size(t):
    t = t.strip().upper()
    m = RECT.match(t)
    if m:
        return f'{m.group(1)}X{m.group(2)}'
    m = ROUND.match(t)
    if m:
        return f'{m.group(1)}"'
    return None


def completed_tag_sizes(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # find header
    hdr = None
    for r in rows:
        u = [str(c).strip().upper() if c else '' for c in r]
        if 'TAG' in u and ('NECK SIZE' in u or 'NECK' in ' '.join(u)):
            hdr = u
            break
    if not hdr:
        return {}
    ti = hdr.index('TAG')
    ni = [i for i, h in enumerate(hdr) if 'NECK' in h]
    ni = ni[0] if ni else ti + 1
    out = defaultdict(set)
    for r in rows:
        c = ['' if x is None else str(x).strip() for x in r]
        if len(c) > max(ti, ni) and c[ti] and c[ti].upper() != 'TAG' and 'Total' not in c[ti]:
            if c[ni]:
                out[c[ti].upper().replace('-', '')].add(c[ni].upper().replace(' ', ''))
    return out


def _expected_shape(type_str):
    """From the schedule type/description, the neck shape we expect.
    PLAQUE FACE / diffuser / ceiling -> round; sidewall / duct / grille -> rect."""
    u = (type_str or '').upper()
    if any(k in u for k in ('SIDEWALL', 'SIDE WALL', 'DUCT', 'GRILLE', 'REGISTER', 'LINEAR')):
        return 'rect'
    if any(k in u for k in ('PLAQUE', 'DIFFUSER', 'CEILING', 'ROUND')):
        return 'round'
    return None


def _shape_of(size_str):
    return 'rect' if 'X' in size_str.upper() else 'round'


def run(pdf, det_json, truth_xlsx, label):
    print(f'\n===== {label} =====')
    d = fitz.open(pdf)
    det = json.load(open(det_json, encoding='utf-8'))
    truth = completed_tag_sizes(truth_xlsx)
    # tag -> expected neck shape, from the parsed schedule
    shape_by_tag = {}
    vg = glob.glob(str(det_json).rsplit('_detections.json', 1)[0] + '_variables.json')
    if vg:
        for v in json.load(open(vg[0], encoding='utf-8')):
            t = v.get('tag')
            if not t:
                continue
            props = v.get('properties') or {}
            typ = ''
            for k, val in props.items():
                if any(w in k.upper() for w in ('TYPE', 'DESCRIPTION', 'SERVICE')):
                    typ = str(val)
                    break
            shape_by_tag[t.upper().replace('-', '')] = _expected_shape(typ)
    tag_sizes = defaultdict(list)
    covered = total = 0
    for pk, dets in det.get('pages', {}).items():
        pidx = int(pk)
        if pidx >= d.page_count:
            continue
        page = d[pidx]
        # Detection px are in DISPLAY (rotated) space; get_text words are in
        # MEDIABOX space. Bridge with the page's derotation matrix (same fix the
        # Bluebeam stamp writer uses), then match nearest size callout.
        mat = fitz.Matrix(page.derotation_matrix)
        words = page.get_text('words')
        toks = []
        for w in words:
            s = parse_size(w[4])
            if s:
                toks.append(((( w[0] + w[2]) / 2, (w[1] + w[3]) / 2), s))
        for x in dets:
            if not str(x.get('cls', '')).startswith('AD'):
                continue
            total += 1
            _c = fitz.Point((x['x1'] + x['x2']) / 2 * DPI_TO_PT,
                            (x['y1'] + x['y2']) / 2 * DPI_TO_PT) * mat
            cx, cy = _c.x, _c.y
            ctag = (x.get('tag') or '?').upper().replace('-', '')
            want = shape_by_tag.get(ctag)
            # candidates within proximity, sorted by distance
            cands = sorted(
                ((math.hypot(sx - cx, sy - cy), s) for (sx, sy), s in toks
                 if math.hypot(sx - cx, sy - cy) <= PROX_PT),
                key=lambda t: t[0])
            best = None
            if cands:
                nearest_d, nearest_s = cands[0]
                best = nearest_s
                # Gentle shape tiebreaker: only override the nearest token with a
                # shape-matching one if that match is nearly as close (within a
                # small margin). Avoids dragging a close-correct callout to a far
                # wrong-shape one (which hurt R2's round 12").
                if want and _shape_of(nearest_s) != want:
                    margin = 1.6 * nearest_d + 25
                    sm = next((s for dd, s in cands if _shape_of(s) == want and dd <= margin), None)
                    if sm is not None:
                        best = sm
            if best:
                covered += 1
                tag_sizes[ctag].append(best)
    print(f'AD detections: {total} | got a nearby size: {covered} ({100*covered//total if total else 0}%)')
    print(f'distinct size tokens found on plan: '
          f'{sum(len(parse_size(w[4]) is not None and [1] or []) for p in range(d.page_count) for w in d[p].get_text("words"))}')
    print('per-tag recovered vs completed takeoff:')
    for t in sorted(set(tag_sizes) | set(truth)):
        rec = dict(Counter(tag_sizes.get(t, [])))
        tru = sorted(truth.get(t, []))
        hit = '✓' if rec and any(s.replace('"', '').replace(' ', '') in ''.join(tru).replace('"', '') for s in rec) else ' '
        print(f'  [{hit}] {t:8} recovered={rec}   truth={tru}')


if __name__ == '__main__':
    DL = 'C:/Users/TriuneTakeoff/Downloads'
    base = 'saas/data/jobs'
    pnc_pdf = f'{DL}/06.20.2026 PNC BANK -Medical Center/06.20.2026 PNC BANK -Medical Center/2026.05.15 - PNC TX Hou Medical Center.pdf'
    pnc_det = glob.glob(f'{base}/64843825f0a0/*_detections.json')[0]
    pnc_truth = f'{DL}/06.20.2026 PNC BANK -Medical Center/06.20.2026 PNC BANK -Medical Center/Completed Takeoff/Takeoff_PNC TX Hou Medical Center.xlsx'
    run(pnc_pdf, pnc_det, pnc_truth, 'PNC Medical')
