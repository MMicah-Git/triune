"""Test schedule parser on small projects only (< 10 MB PDFs)."""
import os
from schedule_parser import parse_pdf_schedules
import openpyxl

PROJ = r'C:\Users\JFL\Downloads\Triune\data to train\projects'

# Only small PDFs — parser hangs on big ones
tests = [
    ('04_Flex_230',       r'04_Flex_230\raw\3.11.26 EAG - 9530 Towne Center Drive (Flex 230).pdf',
                          r'04_Flex_230\excel\Takeoff_EAG - 9530 Towne Center Drive (Flex 230).xlsx'),
    ('23_Shamrock',       r'23_Shamrock_Capital\raw\1.7.26 Shamrock Capital TI.pdf',
                          r'23_Shamrock_Capital\excel\Takeoff_Shamrock_Capital_TI_Triune_Haldeman.xlsx'),
    ('01_Flex_200',       r'01_Flex_200_Corridors\raw\3.11.26 EAG - 9530 Towne Center Drive (Flex 200 + Corridors).pdf',
                          r'01_Flex_200_Corridors\excel\Takeoff_EAG - 9530 Towne Center Drive (Flex 200 + Corridors).xlsx'),
    ('02_Flex_210',       r'02_Flex_210\raw\3.11.26 EAG - 9530 Towne Center Drive (Flex 210).pdf',
                          r'02_Flex_210\excel\Takeoff_EAG - 9530 Towne Center Drive (Flex 210).xlsx'),
    ('03_Flex_220',       r'03_Flex_220\raw\3.11.26 EAG - 9530 Towne Center Drive (Flex 220).pdf',
                          r'03_Flex_220\excel\Takeoff_EAG - 9530 Towne Center Drive (Flex 220).xlsx'),
]

print(f'{"Project":<20}{"Size":<10}{"Sched":<8}{"Tags":<8}{"Excel":<8}{"Recall":<10}{"Precision":<10}')
print('-' * 80)

total_matched = 0
total_excel = 0
total_parser = 0

for name, pdf_rel, xlsx_rel in tests:
    pdf = os.path.join(PROJ, pdf_rel)
    xlsx = os.path.join(PROJ, xlsx_rel)

    if not os.path.exists(pdf) or not os.path.exists(xlsx):
        print(f'{name:<20}FILES MISSING')
        continue

    size_mb = os.path.getsize(pdf) / 1024 / 1024
    if size_mb > 10:
        print(f'{name:<20}SKIP ({size_mb:.0f} MB too big)')
        continue

    try:
        schedules, marks, details, legend, summary = parse_pdf_schedules(pdf)
    except Exception as e:
        print(f'{name:<20}PARSER ERROR: {str(e)[:40]}')
        continue

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    excel_tags = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        tag_col = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                for i, c in enumerate(row):
                    if c and str(c).strip().upper() == 'TAG':
                        tag_col = i
                        break
                continue
            if tag_col is not None and row[tag_col]:
                tag = str(row[tag_col]).strip()
                if tag and 'Total' not in tag and tag != '.':
                    excel_tags.add(tag.upper())
        if tag_col is not None:
            break

    parser_set = set(marks)
    matched = excel_tags & parser_set
    missed = excel_tags - parser_set

    recall = len(matched) / max(len(excel_tags), 1) * 100
    precision = len(matched) / max(len(parser_set), 1) * 100

    total_matched += len(matched)
    total_excel += len(excel_tags)
    total_parser += len(parser_set)

    print(f'{name:<20}{size_mb:>4.0f} MB   {len(schedules):<8}{len(marks):<8}{len(excel_tags):<8}{recall:>5.0f}%    {precision:>5.0f}%')
    if missed:
        print(f'                    MISSED: {sorted(missed)}')
    if parser_set - excel_tags:
        extras = sorted(parser_set - excel_tags)
        print(f'                    EXTRA:  {extras[:10]}{"..." if len(extras)>10 else ""}')

print('-' * 80)
if total_excel > 0:
    overall_recall = total_matched / total_excel * 100
    overall_precision = total_matched / max(total_parser, 1) * 100
    print(f'{"OVERALL":<20}{"":<10}{"":<8}{total_parser:<8}{total_excel:<8}{overall_recall:>5.0f}%    {overall_precision:>5.0f}%')
