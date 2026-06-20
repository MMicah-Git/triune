"""
HVAC Takeoff CLI Tool — Demo Version

Takes a blueprint PDF, runs the v8 model, and outputs:
1. Annotated PDF with colored boxes around detected equipment
2. Excel takeoff with counts per equipment type
3. Summary report (printed to console)

Usage:
    python takeoff_cli.py path/to/blueprint.pdf
    python takeoff_cli.py path/to/blueprint.pdf --conf 0.5
    python takeoff_cli.py path/to/blueprint.pdf --output-dir results/
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stdout.reconfigure(line_buffering=True)

import os
import re
import json
import argparse
import time
from pathlib import Path
from collections import defaultdict

import fitz
import cv2
import numpy as np
from collections import OrderedDict

from tag_extractor import summarize_detections_by_tag
from tag_inference import infer_tags
from schedule_parser import parse_pdf_schedules, dump_variables


# ─── CONFIG ───────────────────────────────────────────────────────────────────

# v10 is the production default (v9 retired; v11 rolled back — scored worse than
# v10, see git log). Override with --model.
DEFAULT_MODEL = 'models/hvac_yolov8s_v10.pt'
DPI = 200
TILE_SIZE = 640
TILE_OVERLAP = 160   # match training tiling (train_yolo.py) so seam-straddling symbols aren't missed
NMS_DIST = 50
DEFAULT_CONF = 0.4

COLORS = [
    (0, 200, 0),     # green
    (200, 0, 0),     # blue (BGR)
    (0, 165, 255),   # orange
    (0, 0, 200),     # red
    (255, 200, 0),   # cyan
    (255, 0, 200),   # magenta
    (0, 255, 255),   # yellow
    (180, 180, 0),   # olive
    (128, 0, 128),   # purple
    (0, 128, 128),   # teal
]


# ─── PROJECT INFO (TITLE BLOCK) ───────────────────────────────────────────────

# Inline label-value patterns: value is on same line after LABEL: value
INLINE_PATTERNS = {
    'scale':       [r'\bSCALE\s*[:=]\s*([0-9/"\'\-=\s\.]{1,30}(?:\'|"|FT)[0-9\'\-"=\s\.]{0,20})',
                    r'\bSCALE\s*[:=]\s*(AS\s+NOTED|NTS|N\.T\.S\.|NONE)'],
    'date':        [r'\bDATE\s*[:=]\s*([0-9]{1,2}[\-/\.][0-9]{1,2}[\-/\.][0-9]{2,4})',
                    r'\b(?:ISSUE[D]?|PLOT)\s*DATE\s*[:=]\s*([0-9]{1,2}[\-/\.][0-9]{1,2}[\-/\.][0-9]{2,4})'],
    'sheet':       [r'\bSHEET\s*(?:NO\.?|NUMBER)\s*[:=]\s*([A-Z]{1,3}[\-\.]?[0-9]{1,4}(?:\.[0-9]+)?)',
                    r'\bDRAWING\s*(?:NO\.?|NUMBER)\s*[:=]\s*([A-Z]{1,3}[\-\.]?[0-9]{1,4})'],
    'project':     [r'\bPROJECT\s*(?:NAME)?\s*[:=]\s*([^\n\r]{3,80})',
                    r'\bJOB\s*(?:NAME|NO\.?)\s*[:=]\s*([^\n\r]{3,80})'],
    'engineer':    [r'\bENGINEER\s*(?:OF\s*RECORD)?\s*[:=]\s*([^\n\r]{3,60})',
                    r'\bDESIGNED\s*BY\s*[:=]\s*([^\n\r]{2,40})',
                    r'\bDRAWN\s*BY\s*[:=]\s*([^\n\r]{2,40})'],
    'firm':        [r'\b(?:MECHANICAL\s+ENGINEER|MEP\s+ENGINEER|CONSULTANT)\s*[:=]\s*([^\n\r]{3,80})'],
    'revision':    [r'\bREV(?:ISION)?\s*(?:NO\.?)?\s*[:=]\s*([0-9A-Z]{1,4})\b'],
}

# Bad-phrase prefixes that should never be treated as values
_BAD_VALUE_STARTS = re.compile(
    r'^(?:ISSUE(?:D)?\b|NOT\b|FOR\b|TO\s+BE\b|APPROVED\b|REVIEWED\b|SEE\b|DESCRIPTION\b|NO\.\b|NOTES?\b|BY\b)',
    re.IGNORECASE,
)

# Field-specific value validators (return True if val passes for that key)
_FIELD_VALIDATORS = {
    'scale':    lambda v: bool(re.search(r'[0-9/]["\']|NTS|N\.T\.S|AS\s+NOTED|NONE', v, re.I)),
    'date':     lambda v: bool(re.search(r'[0-9]{1,2}[\-/\.][0-9]{1,2}[\-/\.][0-9]{2,4}|[A-Z][a-z]+\s+\d{1,2}[,\s]+\d{4}', v)),
    'sheet':    lambda v: bool(re.match(r'^[A-Z]{1,3}[\-\.]?[0-9]{1,4}(?:\.[0-9]+)?$', v.strip())),
    'revision': lambda v: bool(re.match(r'^[0-9]{1,3}$|^[A-Z]$', v.strip())),
    'project':  lambda v: (len(v) >= 5 and not re.match(r'^(?:Project|Job|Sheet|Drawing|Date|Scale|Revision)\s*(?:Name|Number|Title|No\.?)?$', v, re.I)),
    'engineer': lambda v: (2 <= len(v) <= 40 and not re.search(r'\d{3,}', v)
                            and not re.match(r'^(?:DRAWN|DESIGNED|CHECKED)\s+BY$', v.strip(), re.I)
                            and bool(re.search(r'[A-Z]', v))  # initials/name are capitalised
                            and v.strip().lower() not in
                            {'of', 'by', 'the', 'and', 'to', 'for', 'a', 'an', 'as', 'no', 'or'}),
    'firm':     lambda v: len(v) >= 4,
}


def _valid_field(key, val):
    v = _FIELD_VALIDATORS.get(key)
    return v(val) if v else True

# Labels that commonly sit in a column with the value on the NEXT visible line
# (typical CAD title blocks: "Project Name" line, then the actual name below)
STACKED_LABELS = {
    'project':  [r'^PROJECT\s*(?:NAME|TITLE)?\s*:?$', r'^JOB\s*(?:NAME|TITLE)?\s*:?$'],
    'date':     [r'^DATE\s*:?$', r'^ISSUE(?:D)?\s*DATE\s*:?$'],
    'scale':    [r'^SCALE\s*:?$'],
    'sheet':    [r'^SHEET\s*(?:NO\.?|NUMBER)?\s*:?$', r'^DRAWING\s*(?:NO\.?|NUMBER)?\s*:?$'],
    'engineer': [r'^(?:DRAWN|DESIGNED|CHECKED)\s*BY\s*:?$'],
}


def _clean(val):
    val = val.strip().strip(':-=').strip()
    val = re.sub(r'\s+', ' ', val)
    return val


def _is_bad_value(val):
    if not val or len(val) < 2:
        return True
    if _BAD_VALUE_STARTS.match(val):
        return True
    return False


# Module-level YOLO cache. The Arq worker (saas/backend/worker.py) calls
# run_with_args() per job and benefits from the cache — only the first job
# pays the ~5 s model-load cost.
_YOLO_CACHE: dict = {}


def _get_yolo_model(model_path):
    """Return a cached YOLO model, loading it if first call."""
    key = str(model_path)
    if key not in _YOLO_CACHE:
        from ultralytics import YOLO
        _YOLO_CACHE[key] = YOLO(model_path)
    return _YOLO_CACHE[key]


def run_with_args(argv: list[str]) -> int:
    """Programmatic entry point — call as if from the command line.

    Used by the warm-model worker to invoke the pipeline in-process and
    re-use the cached YOLO model across jobs. Returns the process exit
    code (0 on success, non-zero on error).
    """
    import sys as _sys
    saved_argv = _sys.argv
    try:
        _sys.argv = ['takeoff_cli.py'] + list(argv)
        try:
            main()
            return 0
        except SystemExit as e:
            return int(e.code) if isinstance(e.code, int) else 1
        except Exception as e:
            print(f'ERROR in run_with_args: {e!r}')
            import traceback
            traceback.print_exc()
            return 1
    finally:
        _sys.argv = saved_argv


def extract_project_info(pdf_path, max_pages=3):
    """Best-effort extraction of title-block info from a PDF."""
    info = {}
    try:
        doc = fitz.open(str(pdf_path))
        chunks = []
        for i in range(min(max_pages, len(doc))):
            try:
                chunks.append(doc[i].get_text("text"))
            except Exception:
                pass
        doc.close()
    except Exception as e:
        return {'_error': str(e)}

    full_text = "\n".join(chunks)
    if not full_text.strip():
        return info

    # Pass 1 — inline LABEL: VALUE patterns
    for key, patterns in INLINE_PATTERNS.items():
        for pat in patterns:
            m = re.search(pat, full_text, re.IGNORECASE)
            if m:
                val = _clean(m.group(1))
                if not _is_bad_value(val) and _valid_field(key, val):
                    info[key] = val
                    break

    # Pass 2 — stacked labels: value is on the next non-empty line
    lines = [ln.strip() for ln in full_text.split('\n')]
    for i, line in enumerate(lines):
        if not line:
            continue
        for key, patterns in STACKED_LABELS.items():
            if key in info:
                continue
            for pat in patterns:
                if re.match(pat, line, re.IGNORECASE):
                    # Next non-empty line within 3 lines = candidate value
                    for j in range(i + 1, min(i + 4, len(lines))):
                        cand = lines[j]
                        if not cand or len(cand) >= 80:
                            continue
                        if _is_bad_value(cand):
                            continue
                        # Skip if it's another label (ends with colon, or is a known label word)
                        if re.match(r'^(?:Project|Job|Sheet|Drawing|Date|Scale|Revision|Description|Drawn|Designed|Checked|By|No\.?|Number|Name|Title|Tel|Phone|Fax|Email)\s*(?:Name|Number|Title|By|No\.?|:)?\s*:?$', cand, re.I):
                            continue
                        if re.match(r'^[A-Z][A-Z\s]{2,30}:?$', cand):
                            continue
                        clean_cand = _clean(cand)
                        if not _valid_field(key, clean_cand):
                            continue
                        info[key] = clean_cand
                        break
                    break

    # Pass 3 — firm name heuristic (ALL CAPS + engineering/consulting suffix)
    if 'firm' not in info:
        firm_re = re.compile(
            r'(?m)^([A-Z][A-Z&\.\-\' ,]{2,60}\s+(?:ENGINEERING|ENGINEERS|CONSULTANTS?|ASSOCIATES|DESIGN(?:S)?|GROUP|ARCHITECTS?)(?:\s*,?\s*(?:INC|LLC|LLP|P\.?C\.?|PLLC)\.?)?)\s*$'
        )
        for m in firm_re.finditer(full_text):
            cand = _clean(m.group(1))
            if not _is_bad_value(cand):
                info['firm'] = cand
                break

    # Pass 4 — address (street line). Label-aware to avoid grabbing random "9530 Towne..." twice.
    if 'address' not in info:
        addr_re = re.compile(
            r'\b([0-9]{2,5}\s+[A-Z][A-Za-z0-9\.\'\-]+(?:\s+[A-Z][A-Za-z0-9\.\'\-]+){1,5}\s+(?:ST|STREET|AVE|AVENUE|RD|ROAD|BLVD|DR|DRIVE|WAY|LN|LANE|CT|COURT|PL|PLACE|CIR|CIRCLE|CTR|CENTER|CENTRE)\.?)\b'
        )
        m = addr_re.search(full_text)
        if m:
            info['address'] = _clean(m.group(1))

    # Pass 5 — spatial label/value pairing for CAD title blocks (Gensler etc.)
    try:
        spatial = extract_project_info_spatial(pdf_path, max_pages=max_pages)
        for k, v in spatial.items():
            if k not in info or len(info.get(k, '')) < 2:
                info[k] = v
    except Exception:
        pass

    return info


# Labels used by the spatial extractor. Each entry: (key, list of label-text variants).
# Match is case-insensitive, exact-text after stripping trailing colon.
SPATIAL_LABELS = [
    ('project',     ['Project Name', 'Job Name', 'Job Title', 'Project']),
    ('project_no',  ['Project Number', 'Project No.', 'Project No', 'Job Number', 'Job No.', 'Job No']),
    ('description', ['Description', 'Sheet Title', 'Drawing Title']),
    ('sheet',       ['Sheet Number', 'Sheet No.', 'Sheet No', 'Drawing Number', 'Drawing No.']),
    ('scale',       ['Scale']),
    ('date',        ['Date', 'Issue Date', 'Plot Date']),
    ('engineer',    ['Drawn By', 'Designed By', 'Checked By', 'Engineer']),
    ('firm',        ['Architect', 'Mechanical Engineer', 'MEP Engineer', 'Consultant']),
]

_LABEL_TEXTS_LC = {v.lower().rstrip(':') for _, vs in SPATIAL_LABELS for v in vs}


def _spans_in_display_space(page):
    """Return list of {bbox, cx, cy, w, h, text} spans in display coords."""
    rot = page.rotation
    mb_w, mb_h = float(page.mediabox.width), float(page.mediabox.height)
    out = []
    for block in page.get_text('dict').get('blocks', []):
        if block.get('type') != 0:
            continue
        for line in block.get('lines', []):
            for sp in line.get('spans', []):
                txt = sp.get('text', '').strip()
                if not txt:
                    continue
                x0, y0, x1, y1 = sp['bbox']
                if rot == 270:
                    dx0, dy0, dx1, dy1 = mb_h - y1, x0, mb_h - y0, x1
                elif rot == 90:
                    dx0, dy0, dx1, dy1 = y0, mb_w - x1, y1, mb_w - x0
                elif rot == 180:
                    dx0, dy0, dx1, dy1 = mb_w - x1, mb_h - y1, mb_w - x0, mb_h - y0
                else:
                    dx0, dy0, dx1, dy1 = x0, y0, x1, y1
                out.append({
                    'bbox': (dx0, dy0, dx1, dy1),
                    'cx': (dx0 + dx1) / 2, 'cy': (dy0 + dy1) / 2,
                    'w': dx1 - dx0, 'h': dy1 - dy0,
                    'text': txt,
                })
    return out


def _find_value_for_label(label_span, spans):
    """Look for the value span paired with this label.
    CAD title blocks usually place the value DIRECTLY ABOVE the label
    (small label text under a larger value). Fall back to right-of-label.
    """
    lcx, lcy = label_span['cx'], label_span['cy']
    lh = max(label_span['h'], 6.0)
    lw = max(label_span['w'], 30.0)

    candidates = []
    for v in spans:
        if v is label_span:
            continue
        vt = v['text'].strip()
        if not vt or len(vt) < 1:
            continue
        if vt.lower().rstrip(':') in _LABEL_TEXTS_LC:
            continue
        if re.match(r'^[\W_]+$', vt):
            continue
        dx = v['cx'] - lcx
        dy = v['cy'] - lcy   # negative => above in display
        # ABOVE: CAD title blocks often have values that extend wider than the label.
        # Allow generous horizontal tolerance; weight pick by value height (taller = sheet title font).
        if abs(dx) <= max(lw * 3.0, 120) and -lh * 6 < dy < -lh * 0.2:
            score = abs(dy) - v['h'] * 2.0   # prefer closest-above + tallest value
            candidates.append((score, 0, v))
        # RIGHT-OF: value to the right on same baseline (LABEL: VALUE inline rendered as separate span)
        elif abs(dy) <= lh * 0.6 and 0 < dx < lw * 4:
            candidates.append((dx, 1, v))

    if not candidates:
        return None
    candidates.sort(key=lambda t: (t[1], t[0]))
    return candidates[0][2]['text']


def _find_value_with_height(label_span, spans):
    """Same as _find_value_for_label but returns (text, height) so the caller
    can rank competing label instances by value typography size."""
    lcx, lcy = label_span['cx'], label_span['cy']
    lh = max(label_span['h'], 6.0)
    lw = max(label_span['w'], 30.0)
    candidates = []
    for v in spans:
        if v is label_span:
            continue
        vt = v['text'].strip()
        if not vt:
            continue
        if vt.lower().rstrip(':') in _LABEL_TEXTS_LC:
            continue
        if re.match(r'^[\W_]+$', vt):
            continue
        dx = v['cx'] - lcx
        dy = v['cy'] - lcy
        if abs(dx) <= max(lw * 3.0, 120) and -lh * 6 < dy < -lh * 0.2:
            score = abs(dy) - v['h'] * 2.0
            candidates.append((score, 0, v))
        elif abs(dy) <= lh * 0.6 and 0 < dx < lw * 4:
            candidates.append((dx, 1, v))
    if not candidates:
        return None
    candidates.sort(key=lambda t: (t[1], t[0]))
    best = candidates[0][2]
    return best['text'], best['h']


def extract_project_info_spatial(pdf_path, max_pages=3):
    """Bbox-aware title-block extraction. Pairs each known label with the
    nearest value above it (CAD style) or to its right (form style)."""
    info = {}
    try:
        doc = fitz.open(str(pdf_path))
    except Exception:
        return info

    # Title-block fields live only on the cover sheet. After page 1, we'd be
    # picking up per-drawing detail callouts (e.g. SCALE: NONE under each detail
    # box) which is not the project scale.
    TITLEBLOCK_KEYS = {'project', 'project_no', 'description', 'sheet',
                        'scale', 'date', 'engineer', 'firm'}

    for page_idx in range(min(max_pages, len(doc))):
        try:
            page = doc[page_idx]
            spans = _spans_in_display_space(page)
        except Exception:
            continue
        if not spans:
            continue

        # Index labels by normalized exact text
        label_lc_to_key = {}
        for key, variants in SPATIAL_LABELS:
            for v in variants:
                label_lc_to_key.setdefault(v.lower(), key)

        # Collect ALL (label, value, value_height) candidates per key, pick best
        per_key = {}
        for sp in spans:
            t_norm = sp['text'].strip().rstrip(':').lower()
            key = label_lc_to_key.get(t_norm)
            if not key or key in info:
                continue
            # Only accept title-block fields on the cover sheet (page 0).
            # Otherwise we pick up "SCALE: NONE" stamps printed under each
            # detail drawing on later pages.
            if page_idx > 0 and key in TITLEBLOCK_KEYS:
                continue
            pair = _find_value_with_height(sp, spans)
            if not pair:
                continue
            val, vh = pair
            val = _clean(val)
            if _is_bad_value(val):
                continue
            if key in _FIELD_VALIDATORS and not _valid_field(key, val):
                continue
            # Keep the candidate with the tallest value text (real sheet titles
            # are big; column-header "Description" values are small).
            cur = per_key.get(key)
            if cur is None or vh > cur[1]:
                per_key[key] = (val, vh)

        for key, (val, _) in per_key.items():
            if key not in info:
                info[key] = val

    doc.close()
    return info


def print_project_info(info):
    """Pretty-print project info block."""
    if not info or (len(info) == 1 and '_error' in info):
        print(
            "WARNING: title block not parsed.\n"
            "  Most likely cause: non-English labels (FR/ES/etc) or title-block text\n"
            "  embedded as outlined CAD paths rather than searchable text. Pipeline\n"
            "  will continue — downstream stages don't depend on title-block info."
        )
        return
    label_map = [
        ('project',     'Project'),
        ('project_no',  'Project No.'),
        ('description', 'Sheet Title'),
        ('firm',        'Firm'),
        ('engineer',    'Engineer'),
        ('address',     'Address'),
        ('sheet',       'Sheet'),
        ('scale',       'Scale'),
        ('date',        'Date'),
        ('revision',    'Revision'),
    ]
    print("Project info (best-effort from title block):")
    for key, label in label_map:
        if key in info:
            print(f"  {label:<10} {info[key]}")


# ─── PDF HANDLING ─────────────────────────────────────────────────────────────

def render_page(pdf_path, page_idx, dpi=DPI):
    """Render a PDF page to BGR numpy array."""
    doc = fitz.open(pdf_path)
    page = doc[page_idx]
    pix = page.get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72))
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:
        img = cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
    elif pix.n == 3:
        img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    rotation = page.rotation
    mb_w, mb_h = page.mediabox.width, page.mediabox.height
    doc.close()
    return img, rotation, mb_w, mb_h


class LazyPageImages:
    """Dict-like view over rendered page images that renders on demand and keeps
    at most ``cache_size`` pages resident.

    Holding every full-resolution render at once OOM-kills the process on large
    E-size sets (a 36"x24" sheet at 200 DPI is ~100 MB; 16 floor plans is ~1.6 GB).
    Tag inference only ever needs one page image at a time (Level 2b crops bubbles
    page-by-page), so we re-render on access instead of accumulating. Supports the
    access patterns tag_inference uses: ``idx in imgs``, ``imgs[idx]``,
    ``imgs.items()``, and truthiness.
    """

    def __init__(self, pdf_path, page_indices, cache_size=2, dpi=DPI):
        self.pdf_path = str(pdf_path)
        self._keys = list(page_indices)
        self.cache_size = max(1, cache_size)
        self.dpi = dpi
        self._cache = OrderedDict()

    def __bool__(self):
        return bool(self._keys)

    def __len__(self):
        return len(self._keys)

    def __contains__(self, k):
        return k in self._keys

    def __getitem__(self, k):
        if k not in self._keys:
            raise KeyError(k)
        if k in self._cache:
            self._cache.move_to_end(k)
            return self._cache[k]
        img, _rot, _w, _h = render_page(self.pdf_path, k, dpi=self.dpi)
        self._cache[k] = img
        self._cache.move_to_end(k)
        while len(self._cache) > self.cache_size:
            self._cache.popitem(last=False)
        return self._cache[k]

    def keys(self):
        return list(self._keys)

    def items(self):
        for k in self._keys:
            yield k, self[k]


def display_to_annot(dx, dy, rot, mb_w, mb_h):
    """Convert display pixel coords back to annotation (mediabox) coords for adding annotations."""
    # display image is rendered AFTER rotation, so display coords need to be inverted
    if rot == 270:
        # Display: (display_w, display_h) = (mb_h, mb_w)
        # Forward: dx = ay, dy = mb_w - ax => ax = mb_w - dy, ay = dx
        return mb_w - dy, dx
    elif rot == 90:
        return dy, mb_h - dx
    elif rot == 180:
        return mb_w - dx, mb_h - dy
    return dx, dy


# ─── INFERENCE ────────────────────────────────────────────────────────────────

def run_inference(model, img, conf=DEFAULT_CONF):
    """Tile image, run YOLO, return deduplicated detections.

    Two-stage confidence filtering:
      1. YOLO uses a permissive floor (`conf`, usually 0.25) to keep more
         candidates flowing through the pipeline.
      2. After NMS, per-class thresholds from class_thresholds.py knock
         out detections below each class's tuned threshold.
    """
    # Use a permissive floor at the model so we can apply per-class thresholds
    # downstream. Cap at the lowest per-class threshold (~0.25) so we don't
    # waste cycles on extremely-low-confidence noise.
    model_floor = min(conf, 0.25)

    h, w = img.shape[:2]
    step = TILE_SIZE - TILE_OVERLAP

    # Build tile list, skip empty tiles
    tiles = []
    for y in range(0, h, step):
        for x in range(0, w, step):
            xe, ye = min(x + TILE_SIZE, w), min(y + TILE_SIZE, h)
            xs, ys = max(0, xe - TILE_SIZE), max(0, ye - TILE_SIZE)
            tile = img[ys:ye, xs:xe]
            gray = cv2.cvtColor(tile, cv2.COLOR_BGR2GRAY)
            if (gray < 200).mean() < 0.001:   # skip only near-blank tiles (was 0.005 — dropped sparse light-line CAD)
                continue
            if tile.shape[0] < TILE_SIZE or tile.shape[1] < TILE_SIZE:
                p = np.ones((TILE_SIZE, TILE_SIZE, 3), dtype=np.uint8) * 255
                p[:tile.shape[0], :tile.shape[1]] = tile
                tile = p
            tiles.append((tile, xs, ys))

    if not tiles:
        return []

    # Inference
    dets = []
    for tile, xs, ys in tiles:
        results = model.predict(tile, conf=model_floor, verbose=False)
        for r in results:
            for box in r.boxes:
                x1, y1, x2, y2 = box.xyxy[0].tolist()
                dets.append({
                    'cls': model.names[int(box.cls[0])],
                    'conf': float(box.conf[0]),
                    'cx': (x1 + x2) / 2 + xs,
                    'cy': (y1 + y2) / 2 + ys,
                    'x1': x1 + xs,
                    'y1': y1 + ys,
                    'x2': x2 + xs,
                    'y2': y2 + ys,
                })

    # Page-level NMS — replaces the previous center-distance method.
    nms_dets = _page_level_nms(dets)

    # Per-class confidence threshold filter.
    try:
        from class_thresholds import filter_by_class_threshold
        filtered, drops = filter_by_class_threshold(nms_dets, default=conf)
        if drops:
            drops_str = ', '.join(f'{cls}:{n}' for cls, n in sorted(drops.items(), key=lambda x: -x[1])[:4])
            print(f'      [class-threshold drops: {drops_str}]', flush=True)
        return filtered
    except Exception as e:
        print(f'      (class_thresholds unavailable: {e}; using uniform conf)', flush=True)
        return [d for d in nms_dets if d.get('conf', 1.0) >= conf]


# ─── PAGE-LEVEL NMS ───────────────────────────────────────────────────────────

# IoU threshold for within-class duplicate suppression.
NMS_IOU_SAME = 0.45
# Higher threshold for cross-class — must be very high overlap before we assume
# it's the same physical object (otherwise we'd merge legitimate adjacent items).
NMS_IOU_CROSS = 0.70


def _iou_xyxy(a, b):
    ix1 = max(a['x1'], b['x1']); iy1 = max(a['y1'], b['y1'])
    ix2 = min(a['x2'], b['x2']); iy2 = min(a['y2'], b['y2'])
    iw = max(0.0, ix2 - ix1); ih = max(0.0, iy2 - iy1)
    inter = iw * ih
    aw = a['x2'] - a['x1']; ah = a['y2'] - a['y1']
    bw = b['x2'] - b['x1']; bh = b['y2'] - b['y1']
    union = max(1e-9, aw * ah + bw * bh - inter)
    return inter / union


def _page_level_nms(dets, iou_same=NMS_IOU_SAME, iou_cross=NMS_IOU_CROSS):
    """Class-aware IoU NMS + cross-class extreme-overlap pass.

    Returns deduplicated detection list (same dict shape as input).
    """
    if not dets:
        return []

    # Pass 1: per-class greedy NMS
    from collections import defaultdict
    by_cls = defaultdict(list)
    for d in dets:
        by_cls[d['cls']].append(d)

    kept = []
    for cls, arr in by_cls.items():
        arr = sorted(arr, key=lambda d: -d['conf'])
        suppressed = [False] * len(arr)
        for i, di in enumerate(arr):
            if suppressed[i]:
                continue
            kept.append(di)
            for j in range(i + 1, len(arr)):
                if suppressed[j]:
                    continue
                if _iou_xyxy(di, arr[j]) >= iou_same:
                    suppressed[j] = True

    # Pass 2: cross-class extreme overlap — keep the highest-conf box only
    kept_sorted = sorted(kept, key=lambda d: -d['conf'])
    final = []
    suppressed = [False] * len(kept_sorted)
    for i, di in enumerate(kept_sorted):
        if suppressed[i]:
            continue
        final.append(di)
        for j in range(i + 1, len(kept_sorted)):
            if suppressed[j]:
                continue
            dj = kept_sorted[j]
            if di['cls'] == dj['cls']:
                continue  # already handled in pass 1
            if _iou_xyxy(di, dj) >= iou_cross:
                suppressed[j] = True

    return final


# ─── ANNOTATE PDF ─────────────────────────────────────────────────────────────

def annotate_pdf(input_pdf, output_pdf, detections_per_page):
    """
    Add colored rectangles to the PDF for each detection.
    detections_per_page: dict of page_idx -> list of detections
    """
    doc = fitz.open(input_pdf)

    for page_idx, dets in detections_per_page.items():
        page = doc[page_idx]
        rot = page.rotation
        mb_w, mb_h = page.mediabox.width, page.mediabox.height
        scale = DPI / 72  # pixels per PDF point

        for d in dets:
            # Convert pixel coords back to PDF points
            px_x1, px_y1 = d['x1'], d['y1']
            px_x2, px_y2 = d['x2'], d['y2']

            # Pixel → display PDF points
            disp_x1 = px_x1 / scale
            disp_y1 = px_y1 / scale
            disp_x2 = px_x2 / scale
            disp_y2 = px_y2 / scale

            # Display PDF points → annotation (mediabox) coords
            ann_x1, ann_y1 = display_to_annot(disp_x1, disp_y1, rot, mb_w, mb_h)
            ann_x2, ann_y2 = display_to_annot(disp_x2, disp_y2, rot, mb_w, mb_h)

            # Make sure rect is in proper order
            rect = fitz.Rect(
                min(ann_x1, ann_x2),
                min(ann_y1, ann_y2),
                max(ann_x1, ann_x2),
                max(ann_y1, ann_y2)
            )

            # Color (RGB 0-1) by class
            color_idx = hash(d['cls']) % len(COLORS)
            bgr = COLORS[color_idx]
            rgb = (bgr[2] / 255, bgr[1] / 255, bgr[0] / 255)

            annot = page.add_rect_annot(rect)
            annot.set_colors(stroke=rgb)
            annot.set_border(width=2)
            annot.set_info(content=f"{d['cls']} ({d['conf']:.0%})")
            annot.update()

    doc.save(output_pdf)
    doc.close()


# ─── EXCEL OUTPUT ─────────────────────────────────────────────────────────────

def _prop(details, keywords):
    """Tolerant property lookup: any key containing any keyword (case-insensitive)."""
    if not details:
        return ''
    kw_upper = [k.upper() for k in keywords]
    for k, v in details.items():
        k_norm = ' '.join(str(k).upper().split())
        for kw in kw_upper:
            if kw in k_norm:
                return str(v)
    return ''


def write_excel(output_path, detections_per_page, project_name, schedule_details=None,
                variables=None, reconciliation=None):
    """Write Excel takeoff matching team's format.

    If `variables` (rich per-tag schedule data with inferred YOLO class) is
    provided, emit one row PER SCHEDULE TAG even when the AI couldn't
    assign individual detections to specific tags. This produces a
    "schedule-template" Excel the estimator can fill quantities into,
    matching the team's per-tag row format.

    If `reconciliation` (validation_engine.reconcile output) is provided, a
    separate 'Reconciliation' sheet is added with the schedule-vs-detection
    verdicts. The team's 'Triune Takeoff' sheet format is left untouched.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed, skipping Excel output. Install with: pip install openpyxl")
        return False

    if schedule_details is None:
        schedule_details = {}
    if variables is None:
        variables = []

    wb = openpyxl.Workbook()

    # Sheet 1: Triune Takeoff (matches team's format)
    ws = wb.active
    ws.title = 'Triune Takeoff'

    # Header
    ws['A1'] = f'HVAC Takeoff: {project_name}'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')

    # Team's exact headers
    HEADERS = ['PRODUCT', 'BRAND', 'MODEL', 'QTY', 'TAG', 'NECK SIZE',
               'MODULE SIZE', 'DUCT SIZE', 'TYPE', 'MOUNTING', 'REMARK']
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='DDDDDD')

    # Group detections by (class, tag) and fill in schedule details
    grouped = defaultdict(lambda: {'count': 0, 'pages': set()})
    for page_idx, dets in detections_per_page.items():
        for d in dets:
            cls = d['cls']
            tag = d.get('tag') or ''
            key = (cls, tag)
            grouped[key]['count'] += 1
            grouped[key]['pages'].add(page_idx + 1)

    # NEW: If we have rich schedule variables, pre-seed grouped with one
    # entry per schedule tag (count=0) so the Excel emits a row even when
    # the AI couldn't tag specific detections. Match the team's format.
    untagged_by_class = defaultdict(int)
    for (cls, tag), data in grouped.items():
        if not tag:
            untagged_by_class[cls] += data['count']

    # Pre-process variables: drop non-equipment tags (schedule header
    # abbreviations like PH=Phase, V=Voltage that the parser sometimes
    # ingests as tags) and dedupe (tag, cls) where the same tag appears
    # under multiple classes — keep the entry whose class matches the
    # tag-prefix-derived class.
    try:
        from tag_inference import _infer_class_from_tag, TAG_PREFIX_CLASS
    except Exception:
        _infer_class_from_tag = lambda _t: None
        TAG_PREFIX_CLASS = {}

    # Pass 1: bucket variables by tag
    by_tag = defaultdict(list)
    for v in variables:
        t = (v.get('tag') or '').upper()
        if not t:
            continue
        by_tag[t].append(v)

    # Pass 2: for each tag, pick the canonical (tag, class) row
    canonical_pairs = []   # list of (cls, tag) to seed grouped with
    for t, vs in by_tag.items():
        prefix_class = _infer_class_from_tag(t)
        # Drop non-equipment tags: no recognized prefix AND no class on any entry
        if not prefix_class and not any(v.get('inferred_yolo_class') for v in vs):
            continue
        # Pick the variable whose class matches the prefix-derived class;
        # fall back to the first non-empty class; finally to whatever's first.
        chosen = None
        for v in vs:
            if v.get('inferred_yolo_class') == prefix_class:
                chosen = v
                break
        if chosen is None:
            for v in vs:
                if v.get('inferred_yolo_class'):
                    chosen = v
                    break
        if chosen is None:
            chosen = vs[0]
        cls = chosen.get('inferred_yolo_class') or prefix_class or ''
        if not cls:
            continue
        canonical_pairs.append(((cls, t), chosen))

    for (cls, tag), chosen_v in canonical_pairs:
        key = (cls, tag)
        if key not in grouped:
            grouped[key] = {'count': 0, 'pages': set(), '_schedule_only': True}

    # NEW: For classes with N untagged detections and K scheduled tags,
    # add a footer note row so the estimator knows what to budget the qtys to.
    # We collect these and emit them after the per-tag rows below.
    untagged_footer_rows = []
    for cls, n in untagged_by_class.items():
        scheduled_tags_for_class = [v.get('tag') for v in variables
                                   if (v.get('inferred_yolo_class') or '') == cls]
        if n > 0 and scheduled_tags_for_class:
            untagged_footer_rows.append({
                'class': cls,
                'untagged_count': n,
                'available_tags': scheduled_tags_for_class,
            })

    row = 4
    total = 0
    current_cls = None
    # NEW: hoist tag-property lookup to use variables data if richer than
    # schedule_details. variables has the inferred_yolo_class which we want.
    variables_by_tag = {(v.get('tag') or '').upper(): v for v in variables if v.get('tag')}

    for (cls, tag), data in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1])):
        # Skip untagged class-only rows when we have schedule rows for that class
        # (the schedule rows are more useful — they have specs).
        if not tag and any(k[0] == cls and k[1] for k in grouped):
            continue

        # Lookup schedule details for this tag (tolerant of header variations)
        details = schedule_details.get(tag, {})
        # ALSO check variables — same info but might be richer
        if not details and tag.upper() in variables_by_tag:
            details = variables_by_tag[tag.upper()].get('properties') or {}
        # Prefer a combined column like "MANUFACTURER & MODEL" or "MAKE / MODEL".
        # Split on " / " if present (MAKE/MODEL convention), else first space.
        brand_model = _prop(details, ['MANUFACTURER & MODEL', 'MAKE / MODEL', 'MAKE/MODEL'])
        if brand_model and ' / ' in brand_model:
            brand, model = brand_model.split(' / ', 1)
        elif brand_model and ' ' in brand_model:
            brand, model = brand_model.split(' ', 1)
        elif brand_model:
            brand, model = brand_model, ''
        else:
            brand = _prop(details, ['MANUFACTURER', 'BRAND', 'MAKE'])
            model = _prop(details, ['MODEL NUMBER', 'MODEL'])
        neck_size = _prop(details, ['NECK', 'SIZE (NECK)', 'SIZE'])
        etype = _prop(details, ['SERVICE', 'TYPE', 'DESCRIPTION'])
        mounting = _prop(details, ['MOUNTING', 'MOUNT'])

        # When QTY would be 0 but the tag exists in the schedule, default to
        # 1 (assumed-from-schedule) and flag the row for the estimator to
        # visually verify. The schedule is the source of truth for equipment
        # existence; the plan tells us where. If the plan doesn't reveal it
        # (different convention, missing page, AI miss), we still credit
        # the schedule entry. Mutate data['count'] so per-class subtotals
        # and GRAND TOTAL stay consistent with the displayed QTY.
        is_schedule_only = (data['count'] == 0 and tag and
                           tag.upper() in variables_by_tag)
        if is_schedule_only:
            data['count'] = 1
            qty_remark = '⚠ from schedule — verify on plan'
        else:
            qty_remark = ''
        qty = data['count']

        # PRODUCT column: the team writes a descriptive product name (e.g.
        # "LAY-IN", "EGGCRATE RET. GRILLE", "LOUVERED FACE SUPPLY DIFFUSER"),
        # which lives in the schedule's TYPE/DESCRIPTION (already in `etype`).
        # YOLO can't tell those air-device subtypes apart (it dumps them into
        # AD-GRD), so for air devices use the schedule description when it reads
        # like one; otherwise keep the class. Non-air-device rows are unchanged.
        product = cls
        if cls.startswith('AD-') and etype:
            et = etype.strip()
            # Use the schedule's descriptive type as the product name when it
            # reads like one — either a known descriptor word OR a multi-word
            # phrase ("PERFORATED FACE", "DOUBLE DEFLECTION"). Avoids bare
            # manufacturer model codes (single tokens like OMNI / PCS / 50F).
            is_descriptor = bool(re.search(
                r'DIFFUSER|GRILLE|REGISTER|LINEAR|LAY-?IN|EGGCRATE|LOUVER|SLOT|'
                r'SUPPLY|RETURN|EXHAUST|T-?BAR|CEILING|PERFORATED|DEFLECTION|'
                r'FACE|PLAQUE|DRUM|NOZZLE|ADJUSTABLE|FIXED|SIDEWALL', et, re.I))
            if is_descriptor or ' ' in et:
                product = et
        ws.cell(row=row, column=1, value=product if product != current_cls else '')
        current_cls = product
        ws.cell(row=row, column=2, value=brand)
        ws.cell(row=row, column=3, value=model)
        qty_cell = ws.cell(row=row, column=4, value=qty)
        if is_schedule_only:
            qty_cell.fill = PatternFill('solid', fgColor='FFF2CC')  # light yellow
            qty_cell.font = Font(italic=True)
        ws.cell(row=row, column=5, value=tag)
        ws.cell(row=row, column=6, value=neck_size)
        ws.cell(row=row, column=9, value=etype)
        ws.cell(row=row, column=10, value=mounting)
        page_list = ', '.join(str(p) for p in sorted(data['pages']))
        remark = f"Pages: {page_list}" if page_list else ''
        if qty_remark:
            remark = f'{qty_remark}  ·  {remark}' if remark else qty_remark
        ws.cell(row=row, column=11, value=remark)
        total += qty
        row += 1

    # Product totals — skip None/empty class buckets (PH, V style schedule
    # rows that have no equipment class). The per-class sum already includes
    # untagged detections (the (cls, '') key in grouped), so don't add
    # untagged_by_class again — that was the prior double-count.
    per_class_totals = {}
    for cls in sorted(set(c for c, t in grouped.keys() if c)):
        cls_total = sum(d['count'] for (c, t), d in grouped.items() if c == cls)
        per_class_totals[cls] = cls_total
        ws.cell(row=row, column=1, value=f'{cls} Total').font = Font(bold=True)
        ws.cell(row=row, column=4, value=cls_total).font = Font(bold=True)
        row += 1

    # NEW: Footer notes for classes with untagged detections
    if untagged_footer_rows:
        row += 1
        note_fill = PatternFill('solid', fgColor='FFF2CC')  # light yellow
        ws.cell(row=row, column=1, value='⚠ AI could not assign these to specific tags:').font = Font(bold=True, italic=True)
        ws.cell(row=row, column=1).fill = note_fill
        row += 1
        for note in untagged_footer_rows:
            ws.cell(row=row, column=1, value=note['class']).fill = note_fill
            ws.cell(row=row, column=4, value=note['untagged_count']).fill = note_fill
            ws.cell(row=row, column=5, value=', '.join(note['available_tags'][:7])).fill = note_fill
            ws.cell(row=row, column=11, value='Estimator: distribute these across tags').fill = note_fill
            row += 1

    # Grand total = sum of per-class subtotals (which already include the
    # untagged review-block counts). Using `total` alone would drop those.
    grand_total = sum(per_class_totals.values())
    row += 1
    ws.cell(row=row, column=1, value='GRAND TOTAL').font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, value=grand_total).font = Font(bold=True, size=12)

    # Column widths
    widths = [30, 15, 15, 8, 18, 12, 12, 12, 25, 12, 25]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w

    # Sheet 2: RawData (every detection, flat) + QA status columns. This is our
    # diagnostic sheet (not the team format), so it's safe to extend + colour.
    ws2 = wb.create_sheet('RawData')
    RAW_HEADERS = HEADERS + ['QA STATUS', 'QA CONF']
    for ci, h in enumerate(RAW_HEADERS, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='DDDDDD')

    # Row fill by agreement-gated status (ship green, review yellow, fix red).
    QA_FILL = {
        'confirmed': PatternFill('solid', fgColor='C6EFCE'),
        'needs_review': PatternFill('solid', fgColor='FFEB9C'),
        'flagged': PatternFill('solid', fgColor='FFC7CE'),
    }
    STATUS_COL, CONF_COL = len(HEADERS) + 1, len(HEADERS) + 2

    row = 2
    for page_idx in sorted(detections_per_page.keys()):
        for d in sorted(detections_per_page[page_idx], key=lambda x: (x['cls'], x.get('tag') or '')):
            tag = d.get('tag') or ''
            details = schedule_details.get(tag, {})
            brand_model = _prop(details, ['MANUFACTURER & MODEL', 'MANUFACTURER'])
            if brand_model and ' ' in brand_model and not _prop(details, ['MODEL']):
                brand, model = brand_model.split(' ', 1)
            else:
                brand = _prop(details, ['MANUFACTURER', 'BRAND'])
                model = _prop(details, ['MODEL'])
            neck_size = _prop(details, ['NECK', 'SIZE (NECK)', 'SIZE'])
            etype = _prop(details, ['SERVICE', 'TYPE', 'DESCRIPTION'])

            ws2.cell(row=row, column=1, value=d['cls'])
            ws2.cell(row=row, column=2, value=brand)
            ws2.cell(row=row, column=3, value=model)
            ws2.cell(row=row, column=4, value=1)
            ws2.cell(row=row, column=5, value=tag)
            ws2.cell(row=row, column=6, value=neck_size)
            ws2.cell(row=row, column=9, value=etype)

            # QA status + confidence, with the whole row tinted by status.
            status = d.get('qa_status')
            qaconf = d.get('qa_confidence')
            ws2.cell(row=row, column=STATUS_COL, value=status)
            if qaconf is not None:
                ws2.cell(row=row, column=CONF_COL, value=qaconf)
            fill = QA_FILL.get(status)
            if fill is not None:
                for ci in range(1, CONF_COL + 1):
                    ws2.cell(row=row, column=ci).fill = fill
            row += 1

    # Legend for the status colours
    row += 1
    ws2.cell(row=row, column=1, value='Legend:').font = Font(bold=True)
    for label, key in [('confirmed — ship', 'confirmed'),
                       ('needs_review', 'needs_review'),
                       ('flagged — fix', 'flagged')]:
        ws2.cell(row=row, column=2, value=label).fill = QA_FILL[key]
        row += 1

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions[chr(64 + STATUS_COL)].width = 14
    ws2.column_dimensions[chr(64 + CONF_COL)].width = 9

    # Sheet 3: Reconciliation (schedule vs detection self-check). Kept on its
    # own sheet so the team's 'Triune Takeoff' format stays byte-identical.
    if reconciliation is not None:
        try:
            from validation_engine import STATUS_FILL, STATUS_LABEL
            ws3 = wb.create_sheet('Reconciliation')
            r = 1
            ws3.cell(row=r, column=1, value=f'Reconciliation — {project_name}').font = Font(size=13, bold=True)
            r += 1
            conf = reconciliation.get('project_confidence')
            conf_disp = f'{conf:.0%}' if conf is not None else 'n/a'
            ws3.cell(row=r, column=1,
                     value=f"Project trust: {reconciliation.get('tier')} ({conf_disp}) — heuristic, not calibrated").font = Font(italic=True)
            r += 2

            rec_headers = ['CLASS', 'SCHEDULED', 'DETECTED', 'DELTA', 'VERDICT', 'MISSING TAGS']
            for ci, h in enumerate(rec_headers, 1):
                cell = ws3.cell(row=r, column=ci, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='DDDDDD')
            r += 1

            order = {'under': 0, 'over': 1, 'orphan_class': 2, 'match': 3, 'info': 4}
            for row_data in sorted(reconciliation.get('classes', []),
                                   key=lambda x: (order.get(x['status'], 9), x['class'])):
                status = row_data['status']
                fill = PatternFill('solid', fgColor=STATUS_FILL.get(status, 'FFFFFF'))
                vals = [
                    row_data['class'], row_data['expected'], row_data['detected'],
                    row_data['delta'] if status not in ('match', 'info') else '',
                    STATUS_LABEL.get(status, status),
                    ', '.join(row_data['missing_tags'][:12]),
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws3.cell(row=r, column=ci, value=v)
                    if status in ('under', 'over', 'orphan_class'):
                        c.fill = fill
                r += 1

            missing = reconciliation.get('missing_on_plan') or []
            orphans = reconciliation.get('orphan_tags') or []
            if missing:
                r += 1
                ws3.cell(row=r, column=1, value='Scheduled but not found on any plan:').font = Font(bold=True)
                ws3.cell(row=r, column=2, value=', '.join(missing))
                r += 1
            if orphans:
                ws3.cell(row=r, column=1, value='Detected tags not in any schedule:').font = Font(bold=True)
                ws3.cell(row=r, column=2, value=', '.join(orphans))
                r += 1

            for col, w in zip('ABCDEF', [26, 11, 10, 8, 18, 50]):
                ws3.column_dimensions[col].width = w
        except Exception as e:
            print(f"  (reconciliation sheet skipped: {e})")

    wb.save(output_path)
    return True


# ─── MAIN ─────────────────────────────────────────────────────────────────────

# Sheet-title phrases that mark a page as NOT a floor plan. Across the 6
# May-5 LS reviews, ~50 / 61 phantom detections came from pages of these
# types (legends, schedules, details, notes, cover sheets). Skipping them
# at inference time eliminates the largest single phantom source.
NON_PLAN_TITLE_MARKERS = [
    'MECHANICAL LEGEND', 'HVAC LEGEND', 'PLUMBING LEGEND',
    'LEGEND AND ABBREVIATIONS', 'LEGENDS AND SCHEDULES', 'SCHEDULE AND LEGEND',
    'GENERAL NOTES', 'MECHANICAL NOTES', 'HVAC NOTES',
    'MECHANICAL SCHEDULE', 'HVAC SCHEDULE', 'EQUIPMENT SCHEDULE',
    'AIR DEVICE SCHEDULE', 'DIFFUSER SCHEDULE', 'FAN SCHEDULE',
    'MECHANICAL DETAILS', 'HVAC DETAILS', 'TYPICAL DETAILS',
    'PIPING DETAILS', 'INSTALLATION DETAILS',
    'TITLE SHEET', 'COVER SHEET', 'SHEET INDEX', 'DRAWING INDEX',
    'SYMBOLS AND ABBREVIATIONS',
]

PLAN_KEYWORDS = [
    'MECHANICAL PLAN', 'CEILING PLAN', 'HVAC PLAN',
    'VENTILATION PLAN', 'FLOOR PLAN', 'ROOF PLAN',
]


def _is_non_plan_sheet(text_upper):
    """True if page is a legend / schedule / details / notes sheet.

    Match on compound title phrases (e.g., 'MECHANICAL SCHEDULE') rather
    than bare words so floor plans that happen to mention 'SCHEDULE' or
    'LEGEND' in a callout aren't filtered out.
    """
    return any(m in text_upper for m in NON_PLAN_TITLE_MARKERS)


def find_mechanical_pages(pdf_path):
    """
    Pick the pages of a multi-discipline drawing set that should run through
    the takeoff pipeline.

    Strategy (in order):
      1. Sheet-number-based filter (sheet_filter.py) — read the title block,
         classify discipline, drop non-M-series and non-plan sheets.
         This is the primary path; works on any drawing set that has a text
         layer in the title block.
      2. Fallback: legacy keyword heuristic — for PDFs with no detectable
         sheet numbers (e.g. raster-only addenda).
    """
    # Primary: sheet-number-aware filter
    try:
        from sheet_filter import survey_pdf, is_m_series
    except Exception as e:
        print(f"  (sheet_filter unavailable: {e}; falling back to keyword scan)")
        survey_pdf = None  # type: ignore
        is_m_series = None  # type: ignore

    survey = survey_pdf(pdf_path) if survey_pdf else None

    if survey:
        keepers = [s for s in survey if s.is_plan and is_m_series(s.discipline)]
        skipped = [s for s in survey if not (s.is_plan and is_m_series(s.discipline))]

        if keepers:
            # Concise log
            keep_list = ', '.join(
                f"p{s.page_idx+1}={s.sheet_number}" for s in keepers
            )
            print(f"  Sheet filter kept {len(keepers)} M-series plan page(s): {keep_list}")
            if skipped:
                # Group skipped by reason for a tight summary
                from collections import Counter
                reasons = Counter()
                for s in skipped:
                    if s.sheet_number is None:
                        reasons['no sheet number'] += 1
                    elif not is_m_series(s.discipline):
                        reasons[f'{s.discipline}-series'] += 1
                    elif not s.is_plan:
                        reasons['non-plan title (legend/schedule/details/notes)'] += 1
                bits = ', '.join(f'{n} {k}' for k, n in reasons.most_common())
                print(f"  Skipped {len(skipped)} page(s): {bits}")
            return [s.page_idx for s in keepers]

        # Sheet filter ran but found nothing — either non-standard PDF or
        # raster-only mechanical pages. Fall through to keyword scan.
        print("  Sheet filter found no M-series plan pages; falling back to keyword scan.")

    # Fallback: keyword scan
    doc = fitz.open(pdf_path)
    total = doc.page_count
    candidate_pages = []
    skipped_non_plan = []
    for pi in range(total):
        text = doc[pi].get_text().upper()
        if not any(kw in text for kw in PLAN_KEYWORDS):
            continue
        if _is_non_plan_sheet(text):
            skipped_non_plan.append(pi + 1)
            continue
        candidate_pages.append(pi)
    doc.close()

    if skipped_non_plan:
        print(f"  Skipping non-plan sheets: pages {skipped_non_plan}")

    if not candidate_pages:
        return list(range(total))
    return candidate_pages


# Schedule tables carry a tag/mark column plus several property columns. We use
# this signal — not the bare word "SCHEDULE" (which appears in general notes
# prose like "IDENTIFIED IN EQUIPMENT SCHEDULE") — to decide which pages to feed
# to the (slow) pdfplumber table extractor.
_SCHED_TAG_HEADERS = ('MARK', 'TAG', 'UNIT TAG', 'UNIT NO', 'EQUIP', 'DESIGNATION')
_SCHED_PROP_KEYS = ('CFM', 'MODEL', 'MANUFACTURER', 'MBH', 'TONS', 'CAPACITY',
                    'NECK', 'HP', 'VOLT', 'PHASE', 'SIZE', 'SERVICE', 'GPM',
                    'ESP', 'RPM', 'WEIGHT', 'SEER', 'BTU')


def find_schedule_pages(pdf_path):
    """Cheap text-layer scan for pages that actually contain equipment SCHEDULE
    *tables* (a tag/mark column header + several property columns).

    The old code restricted schedule parsing to the M-series *plan* pages, which
    (a) misses dedicated schedule sheets and (b) runs pdfplumber's expensive
    table extraction over dense E-size floor plans (15k+ vector paths each → many
    minutes + heavy memory). Targeting only real schedule-table pages fixes both.
    Returns 0-based page indices (empty list if none look like schedule tables).
    """
    try:
        doc = fitz.open(str(pdf_path))
    except Exception:
        return []
    hits = []
    try:
        for i in range(doc.page_count):
            up = doc[i].get_text('text').upper()
            if not up.strip():
                continue
            if 'SCHEDULE' not in up:
                continue
            if not any(h in up for h in _SCHED_TAG_HEADERS):
                continue
            if sum(1 for k in _SCHED_PROP_KEYS if k in up) >= 4:
                hits.append(i)
    finally:
        doc.close()
    return hits


# Plausible HVAC tag: 1-4 letters, optional dash, 1-3 digits, optional suffix
# letter (A-1, EF-12, CU-6B, VAV12). Used to reject OCR noise tags.
_TAG_RE = re.compile(r'^[A-Z]{1,4}-?\d{1,3}[A-Z]?$')
# Common English/CAD words the OCR fallback tends to scrape from notes &
# the "NOT FOR CONSTRUCTION" watermark and mistake for tags.
_OCR_TAG_STOPWORDS = {
    'AND', 'FOR', 'NOT', 'THE', 'OF', 'OR', 'TO', 'IN', 'ON', 'AT', 'AS',
    'PER', 'SEE', 'ALL', 'NTS', 'TYP', 'REF', 'REV', 'NEW', 'EXIST',
    'ARIZ', 'ZONA', 'US', 'USA', 'IBC', 'IMC', 'IECC',
}


def filter_ocr_variables(ocr_vars):
    """Drop OCR-recovered 'schedule variables' that are clearly noise.

    The raster OCR fallback can scrape note prose and the NOT-FOR-CONSTRUCTION
    watermark into fake tags (e.g. tag='AND' with properties that are full
    sentences, or 'P' -> {'0':'ARIZ','NOT':'ZONA'} from 'ARIZONA'). Keep only
    variables whose tag looks like a real HVAC tag and isn't a stopword.
    Returns (kept, dropped_count).
    """
    kept = []
    dropped = 0
    for v in ocr_vars or []:
        tag = (v.get('tag') or '').strip().upper()
        if not tag or tag in _OCR_TAG_STOPWORDS or not _TAG_RE.match(tag):
            dropped += 1
            continue
        kept.append(v)
    return kept, dropped


def main():
    parser = argparse.ArgumentParser(description='HVAC Takeoff CLI Tool')
    parser.add_argument('pdf', help='Path to blueprint PDF')
    parser.add_argument('--model', default=DEFAULT_MODEL, help='Path to YOLO model')
    parser.add_argument('--conf', type=float, default=DEFAULT_CONF, help='Confidence threshold (0-1)')
    parser.add_argument('--output-dir', default=None, help='Output directory (default: same as PDF)')
    parser.add_argument('--all-pages', action='store_true', help='Process all pages (not just mechanical)')
    parser.add_argument('--pages', type=int, nargs='+', help='Specific page numbers (1-indexed)')
    parser.add_argument('--verify', action='store_true',
                        help='Print full schedule variable dump and exit (no detection run)')
    parser.add_argument('--schedule-only', action='store_true',
                        help='Parse schedule and write variables JSON, skip YOLO detection')
    parser.add_argument('--no-schedule-ocr', dest='no_schedule_ocr', action='store_true',
                        help='Skip the raster OCR schedule fallback. The fallback renders '
                             'E-size pages at 200 DPI + EasyOCR (slow + memory-heavy) and on '
                             'text-readable sets can scrape noise; use this to bypass it.')
    # Future-proofing placeholders — accepted but currently no-op. Reserved for
    # multilingual keyword-set support (see sheet_filter / schedule_parser).
    parser.add_argument('--english-only', action='store_true',
                        help='(reserved) enforce strict English keyword matching')
    parser.add_argument('--lang', default=None,
                        help='(reserved) project language hint (en/fr/es/de). '
                             'Currently no-op; will swap keyword sets in a future build')
    args = parser.parse_args()

    pdf_path = Path(args.pdf).resolve()
    if not pdf_path.exists():
        print(f"ERROR: {pdf_path} not found")
        sys.exit(1)

    if not args.schedule_only and not Path(args.model).exists():
        print(f"ERROR: Model not found: {args.model}")
        sys.exit(1)

    # Output directory
    if args.output_dir:
        out_dir = Path(args.output_dir)
    else:
        out_dir = pdf_path.parent / f"{pdf_path.stem}_takeoff"
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*70}")
    print(f"HVAC TAKEOFF — {pdf_path.name}")
    print(f"{'='*70}")
    print(f"Model:    {args.model}")
    print(f"Conf:     {args.conf}")
    print(f"Output:   {out_dir}")
    print()

    # Extract title-block info (best-effort)
    project_info = extract_project_info(pdf_path)
    print_project_info(project_info)
    print()

    # Run the sheet filter ONCE up-front so we can:
    #   (a) restrict schedule parsing to M-family pages (huge speedup)
    #   (b) reuse the YOLO plan-page list later without re-surveying
    # Skip the filter when the user passed --all-pages or --pages, so those
    # debug flags retain their original semantics.
    cached_m_plan_pages = None
    cached_m_series_pages = None
    if not args.all_pages and not args.pages:
        try:
            from sheet_filter import survey_summary, is_m_series
            print("Detecting M-series pages...")
            summary = survey_summary(pdf_path)
            cached_m_plan_pages = summary['m_plan_pages']
            cached_m_series_pages = summary['m_series_pages']
            survey = summary['survey']
            n_total = len(survey)
            n_plan = len(cached_m_plan_pages)
            n_series = len(cached_m_series_pages)
            print(f"  {n_series} M-series page(s) of {n_total}; {n_plan} are floor plans")
            kept = ', '.join(
                f"p{s.page_idx+1}={s.sheet_number}"
                for s in survey
                if is_m_series(s.discipline)
            )
            if kept:
                print(f"  M-series pages: {kept}")
            print()
        except Exception as e:
            print(f"  (sheet_filter unavailable: {e}; will scan all pages)\n")

    # Parse schedule first — always, even in --schedule-only mode.
    # If we have a cached M-series page list, restrict pdfplumber to those
    # pages only — saves the bulk of schedule-scan time on multi-discipline
    # drawing sets.
    print("Parsing schedule...")
    variables = []
    # Schedules live on schedule SHEETS, not necessarily the plan pages. Target
    # pages that actually contain schedule tables (tag column + property columns)
    # instead of running pdfplumber over every dense E-size plan page. Fall back
    # to the M-series page list only if no schedule-table page is detected.
    schedule_pages = find_schedule_pages(pdf_path)
    if schedule_pages:
        print(f"  Schedule-table pages: {[p + 1 for p in schedule_pages]}")
    else:
        schedule_pages = cached_m_series_pages
    try:
        schedules, marks, mark_details, legend, sched_summary, variables = parse_pdf_schedules(
            str(pdf_path),
            pages=schedule_pages,
        )
        print(f"  {len(schedules)} schedule table(s), {len(marks)} unique tag(s), "
              f"{len(variables)} variable(s)")
        if marks:
            print(f"  Sample tags: {marks[:10]}{'...' if len(marks) > 10 else ''}")
    except Exception as e:
        print(f"  Schedule parse failed: {e}")
        schedules, marks, mark_details = [], [], {}

    # OCR fallback: when the text-layer parser found no variables (e.g. a
    # CAD-exported PDF with broken/non-extractable font encoding), read the
    # schedule from the rendered pixels instead. The OCR module lives in
    # saas/backend; import it lazily so the CLI has no hard EasyOCR dependency.
    if not variables and not getattr(args, 'no_schedule_ocr', False):
        try:
            import sys as _sys
            _ocr_dir = Path(__file__).resolve().parent / 'saas' / 'backend'
            if _ocr_dir.is_dir() and str(_ocr_dir) not in _sys.path:
                _sys.path.insert(0, str(_ocr_dir))
            from schedule_ocr import extract_all_schedules as _ocr_sched
            # OCR only the pages that look like schedule TABLES, not the dense
            # E-size plan pages (rendering 8 of those at 200 DPI is the slow,
            # ~5 GB memory-heavy step that yields only noise). schedule_pages are
            # 0-based; schedule_ocr expects 1-based page numbers.
            _ocr_src = schedule_pages or cached_m_series_pages
            if _ocr_src:
                ocr_pages = [p + 1 for p in _ocr_src][:8]
            else:
                import fitz as _fitz
                _d = _fitz.open(str(pdf_path)); _n = _d.page_count; _d.close()
                ocr_pages = list(range(1, min(_n, 8) + 1))
            print(f"  Text-layer schedule empty; trying OCR fallback on pages {ocr_pages}...")
            ocr_vars = _ocr_sched(str(pdf_path), ocr_pages, dpi=200)
            ocr_vars, ocr_dropped = filter_ocr_variables(ocr_vars)
            if ocr_dropped:
                print(f"  OCR guard discarded {ocr_dropped} noise 'variable(s)' "
                      f"(non-tag text scraped from notes/watermark)")
            if ocr_vars:
                variables = ocr_vars
                marks = sorted({v.get('tag') for v in variables if v.get('tag')})
                print(f"  OCR recovered {len(variables)} schedule variable(s): "
                      f"{marks[:10]}{'...' if len(marks) > 10 else ''}")
            else:
                print("  OCR fallback recovered 0 usable variables "
                      "(no schedule tables found in this PDF)")
        except Exception as _e:
            print(f"  OCR fallback unavailable/failed: {_e}")

    # Always write variables JSON sidecar
    variables_path = out_dir / f"{pdf_path.stem}_variables.json"
    try:
        with open(variables_path, 'w', encoding='utf-8') as f:
            json.dump(variables, f, indent=2, default=str, ensure_ascii=False)
        print(f"  Wrote {len(variables)} variables to {variables_path.name}")
    except Exception as e:
        print(f"  (JSON sidecar failed: {e})")

    # Project info sidecar
    if project_info and not ('_error' in project_info and len(project_info) == 1):
        info_path = out_dir / f"{pdf_path.stem}_project_info.json"
        try:
            with open(info_path, 'w', encoding='utf-8') as f:
                json.dump(project_info, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    # Verification dump to stdout
    if args.verify or args.schedule_only:
        dump_variables(variables)

    # Early exit if schedule-only
    if args.schedule_only:
        print(f"\n--schedule-only: skipping detection. Output in {out_dir}")
        return

    print()

    # Load model (cached at module level — when this is called repeatedly
    # by the warm-model worker, only the first call pays the load cost).
    print("Loading model...")
    model = _get_yolo_model(args.model)
    print(f"  Loaded with {len(model.names)} equipment classes\n")

    # Determine which pages to process
    doc = fitz.open(pdf_path)
    total_pages = doc.page_count
    doc.close()

    if args.pages:
        pages_to_process = [p - 1 for p in args.pages]
    elif args.all_pages:
        pages_to_process = list(range(total_pages))
    else:
        # Fused page selection (Part 1): ONE per-page verdict that combines the
        # sheet-number read, the content type, and drawing density, then keeps a
        # mechanical page UNLESS the content is confidently a non-drawing sheet
        # (schedule/legend/notes/cover/details). Replaces the old two-stacked
        # filters (sheet_filter THEN a subordinate page_classifier pass), which
        # silently dropped real plans numbered M0.x / M5.x / M8.x and pages whose
        # title block wouldn't OCR, while keeping schedules numbered in the plan
        # range (phantom source). See page_selector.py + ARCHITECTURE_PARTS.md.
        try:
            import page_selector
            verdicts = page_selector.classify_pages(pdf_path)
            pages_to_process = [v['page'] - 1 for v in verdicts if v['is_plan']]
            dropped = [v['page'] for v in verdicts if not v['is_plan']]
            print(f"  Page selection (fused): keep {[p + 1 for p in pages_to_process]}"
                  + (f"; drop {dropped}" if dropped else ""))
            try:
                with open(out_dir / f"{pdf_path.stem}_page_selection.json", 'w',
                          encoding='utf-8') as _f:
                    json.dump(verdicts, _f, indent=2)
            except Exception:
                pass
            if not pages_to_process:
                print("  WARN: fused selector kept 0 pages; falling back to all pages")
                pages_to_process = list(range(total_pages))
        except Exception as _e:
            print(f"  (fused page selection unavailable: {_e}; using fallback)")
            if cached_m_plan_pages:
                pages_to_process = cached_m_plan_pages
            else:
                pages_to_process = find_mechanical_pages(pdf_path) or list(range(total_pages))

    print(f"Processing {len(pages_to_process)} page(s) of {total_pages} total\n")

    # Process each page. We do NOT keep every rendered image in memory — on
    # large E-size sets that peaks at ~1.6 GB and OOM-kills the process. The
    # current page image is detected on, then released; tag inference re-renders
    # pages on demand via LazyPageImages (it only needs one page at a time).
    detections_per_page = {}
    t_start = time.time()
    for page_idx in pages_to_process:
        t0 = time.time()
        print(f"  Page {page_idx+1}: rendering...", end=' ', flush=True)
        try:
            img, rot, mb_w, mb_h = render_page(pdf_path, page_idx)
        except Exception as e:
            print(f"FAILED: {e}")
            continue
        print(f"detecting...", end=' ', flush=True)
        dets = run_inference(model, img, conf=args.conf)
        elapsed = time.time() - t0
        print(f"{len(dets)} found ({elapsed:.0f}s)")
        if dets:
            detections_per_page[page_idx] = dets
        del img   # release the full-res render before the next page

    # Memory-bounded, lazily-rendered image view for tag inference (Level 2b OCR).
    page_images = LazyPageImages(pdf_path, list(detections_per_page.keys()))

    total_elapsed = time.time() - t_start
    print(f"\nDetection complete in {total_elapsed:.0f}s")

    # Tag inference — 3-level system
    if detections_per_page:
        print("\nInferring tags...")
        detections_per_page, tag_stats = infer_tags(
            detections_per_page, schedules, marks, mark_details, str(pdf_path),
            variables=variables, page_images=page_images
        )
        print(f"  Tagged: {tag_stats['tagged']}/{tag_stats['total']} ({tag_stats['tagged_pct']:.0f}%)")
        for ls in tag_stats.get('levels', []):
            if ls.get('tagged', 0) > 0 or ls.get('mapping'):
                print(f"  Level {ls.get('level', '?')}: {ls.get('method', '')} — {ls}")

    # ── Change #1: subtype-from-tag ──────────────────────────────────────────
    # YOLO reliably FINDS air devices but can't distinguish the visually-similar
    # subtypes (T-BAR / SURF / LINEAR, supply/return) — it dumps most into the
    # generic AD-GRD. The subtype isn't in the symbol; it's in the tag's schedule
    # row. So for each tagged air-device detection, override its class with the
    # tag's scheduled subtype. Benchmark (rescore_aliased.py): the model's
    # air-device OBJECT recall is ~84% held-out — this captures that as correct
    # per-subtype output without retraining. Refines WITHIN the air-device family
    # only (never reclassifies across equipment families).
    if variables and detections_per_page:
        from collections import Counter as _C
        _tag_sub = {}
        for v in variables:
            _t, _s = v.get('tag'), v.get('inferred_yolo_class')
            if _t and _s:
                _tag_sub.setdefault(_t, _C())[_s] += 1
        tag_to_subtype = {t: c.most_common(1)[0][0] for t, c in _tag_sub.items()}
        _AD = 'AD-'
        _GENERIC = {'AD-GRD'}   # least-specific air-device class
        refined = _C()
        for dets in detections_per_page.values():
            for d in dets:
                tag = d.get('tag')
                cur = (d.get('cls') or '')
                sub = tag_to_subtype.get(tag)
                # Only refine TO a more specific air-device subtype. Never replace
                # a specific YOLO label (AD-T-BAR SUPPLY) with the generic AD-GRD
                # — the schedule is sometimes generic, and degrading would lose
                # information the detector already had.
                if (tag and sub and cur.startswith(_AD) and sub.startswith(_AD)
                        and sub != cur and sub not in _GENERIC):
                    d['cls_detected'] = cur          # keep original for traceability
                    d['cls'] = sub
                    d['cls_source'] = 'tag-schedule'
                    refined[f'{cur}→{sub}'] += 1
        if refined:
            total = sum(refined.values())
            print(f"  Subtype-from-tag: refined {total} air-device detection(s) "
                  f"to scheduled subtype")
            for k, n in refined.most_common(6):
                print(f"    {n} × {k}")

    # Schedule-conditioned post-filter — drop YOLO predictions whose normalized
    # class doesn't appear in the project's schedule, UNLESS the model is very
    # confident (force-keep threshold). This eliminates rare-class false
    # positives like predicting BACKDRAFT DAMPER on a project that has none.
    if variables and detections_per_page:
        try:
            from class_normalization import normalize_class as _norm
            from collections import Counter as _Counter
            # High-confidence detections survive even if their class isn't in the
            # (often partially-parsed) schedule. Lowered 0.85→0.60: a partial
            # schedule was suppressing valid detections worse than no schedule.
            FORCE_KEEP_CONF = 0.60
            valid_classes = {
                _norm(v.get('inferred_yolo_class') or '')
                for v in variables if v.get('inferred_yolo_class')
            }
            valid_classes.discard('')
            if valid_classes:
                dropped = _Counter()
                kept_pages = {}
                for pno, dets in detections_per_page.items():
                    kept = []
                    for d in dets:
                        d_cls = _norm(d.get('cls') or '')
                        if d_cls in valid_classes or d.get('conf', 0) >= FORCE_KEEP_CONF:
                            kept.append(d)
                        else:
                            dropped[d_cls] += 1
                    kept_pages[pno] = kept
                detections_per_page = kept_pages
                total_dropped = sum(dropped.values())
                if total_dropped > 0:
                    print(f"  Schedule filter dropped {total_dropped} prediction(s) "
                          f"whose class isn't in schedule (valid: {sorted(valid_classes)[:6]}{'…' if len(valid_classes)>6 else ''})")
                    for cls, n in dropped.most_common(6):
                        print(f"    -{n} × {cls}")
        except Exception as e:
            print(f"  (schedule-conditioned filter skipped: {e})")
    print()

    # Aggregate
    total_count = sum(len(d) for d in detections_per_page.values())
    if total_count == 0:
        print("No HVAC equipment detected in the selected pages.")
        sys.exit(0)

    class_counts = defaultdict(int)
    for dets in detections_per_page.values():
        for d in dets:
            class_counts[d['cls']] += 1

    # Print summary
    all_dets = [d for dets in detections_per_page.values() for d in dets]
    tag_summary = summarize_detections_by_tag(all_dets)
    tagged = sum(1 for d in all_dets if d.get('tag'))

    print(f"{'='*75}")
    print(f"TAKEOFF SUMMARY")
    print(f"{'='*75}")
    print(f"  Total equipment detected: {total_count}")
    print(f"  Tagged:                   {tagged} / {total_count} ({tagged/max(total_count,1)*100:.0f}%)")
    print(f"  Pages with equipment:     {len(detections_per_page)}")
    print()
    print(f"  {'Equipment Type':<30} {'Tag':<15} {'Count':>8}")
    print(f"  {'-'*30} {'-'*15} {'-'*8}")
    for row in tag_summary:
        tag_disp = row['tag'] if row['tag'] != '(no-tag)' else '—'
        print(f"  {row['class'][:29]:<30} {tag_disp:<15} {row['count']:>8}")
    print()

    # Schedule ↔ detection reconciliation (the closed-loop self-check).
    # Compares scheduled counts/tags against what YOLO actually detected and
    # surfaces under/over/missing/orphan verdicts + a project trust score.
    reconciliation = None
    try:
        from validation_engine import reconcile, format_report
        reconciliation = reconcile(variables, detections_per_page,
                                   conf_threshold=0.0)
        print(format_report(reconciliation))
        print()
    except Exception as e:
        print(f"  (reconciliation skipped: {e})\n")

    # Agreement-gated QA status — one evidence-carrying LineItem per detection.
    # Annotates each detection in place with qa_status / qa_confidence / qa_flags
    # so the detections.json + Excel surface them. Ship `confirmed`, review rest.
    line_items = None
    try:
        from line_items import build_line_items, format_summary
        line_items = build_line_items(detections_per_page, variables, annotate=True)
        print(format_summary(line_items))
        print()
    except Exception as e:
        print(f"  (QA status skipped: {e})\n")

    # Output files
    annotated_pdf_path = out_dir / f"{pdf_path.stem}_annotated.pdf"
    excel_path = out_dir / f"{pdf_path.stem}_takeoff.xlsx"

    print(f"Writing outputs...")
    print(f"  Annotated PDF:  {annotated_pdf_path}")
    annotate_pdf(str(pdf_path), str(annotated_pdf_path), detections_per_page)

    print(f"  Excel takeoff:  {excel_path}")
    write_excel(str(excel_path), detections_per_page, pdf_path.stem, mark_details,
                variables=variables, reconciliation=reconciliation)

    # Reconciliation sidecars (machine + human readable)
    if reconciliation is not None:
        recon_json_path = out_dir / f"{pdf_path.stem}_reconciliation.json"
        with open(recon_json_path, 'w', encoding='utf-8') as f:
            json.dump(reconciliation, f, indent=2, ensure_ascii=False)
        print(f"  Reconciliation: {recon_json_path}")
        try:
            from validation_engine import format_report as _fmt
            recon_txt_path = out_dir / f"{pdf_path.stem}_reconciliation.txt"
            with open(recon_txt_path, 'w', encoding='utf-8') as f:
                f.write(_fmt(reconciliation) + '\n')
        except Exception:
            pass

    # Line-items sidecar (evidence + QA status per detection)
    if line_items is not None:
        li_path = out_dir / f"{pdf_path.stem}_line_items.json"
        with open(li_path, 'w', encoding='utf-8') as f:
            json.dump(line_items, f, indent=2, ensure_ascii=False)
        print(f"  Line items:     {li_path}")

    detections_json_path = out_dir / f"{pdf_path.stem}_detections.json"
    print(f"  Detections:     {detections_json_path}")
    det_dump = {
        'pdf': str(pdf_path),
        'dpi': DPI,
        'pages': {
            str(page_idx): [
                {
                    'cls': d['cls'],
                    'tag': d.get('tag'),
                    'tag_method': d.get('tag_method'),
                    'conf': d.get('conf'),
                    'qa_status': d.get('qa_status'),
                    'qa_confidence': d.get('qa_confidence'),
                    'qa_flags': d.get('qa_flags'),
                    'x1': d['x1'], 'y1': d['y1'], 'x2': d['x2'], 'y2': d['y2'],
                }
                for d in dets
            ]
            for page_idx, dets in detections_per_page.items()
        },
    }
    with open(detections_json_path, 'w', encoding='utf-8') as f:
        json.dump(det_dump, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*70}")
    print(f"DONE — open {out_dir} to see the results")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
