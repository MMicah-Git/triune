"""benchmark_samples.py — End-to-end benchmark across the team's sample projects.

Each project under SAMPLE_ROOT has:
  Plans_Specs/<plan>.pdf       (input — what we feed the CLI)
  Completed Takeoff/<x>.xlsx   (truth — what the team produced manually)

We run takeoff_cli.py on each plan PDF, then score our generated xlsx against
the team's by per-product QTY overlap (primary) and per-(product, tag) QTY
overlap (secondary).

Outputs:
  benchmark_output/benchmark_results.csv      — one row per project
  benchmark_output/benchmark_per_product.csv  — one row per (project, product)
  benchmark_output/benchmark_summary.md       — human-readable leaderboard
  benchmark_output/<project>_takeoff/         — per-project pipeline outputs

Usage:
  python benchmark_samples.py
  python benchmark_samples.py --projects "Sola Salons" "Krispy Kreme"
  python benchmark_samples.py --cache               # skip projects with existing xlsx
  python benchmark_samples.py --root "/path/to/SAMPLE FILES"
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stdout.reconfigure(line_buffering=True)

import argparse
import csv
import json
import re
import subprocess
import time
from collections import defaultdict
from pathlib import Path


SAMPLE_ROOT = Path(r'C:\Users\JFL\Downloads\SAMPLE FILES 27.04.26\SAMPLE FILES 27.04.26')
TOOL_DIR = Path(__file__).resolve().parent
OUT_ROOT = TOOL_DIR / 'benchmark_output'
MAX_PLAN_MB = 20.0


# ─── Project discovery ─────────────────────────────────────────────────────

def discover_projects(root):
    """Walk root, return list of dicts: {name, plan_pdf, truth_xlsx}.

    Skips KNAPE FILE/ (different structure) and any project missing either
    a Plans_Specs/*.pdf or a Completed Takeoff/*.xlsx.
    """
    projects = []
    for d in sorted(root.iterdir()):
        if not d.is_dir() or d.name == 'KNAPE FILE':
            continue
        plans_dir = d / 'Plans_Specs'
        truth_dir = d / 'Completed Takeoff'
        if not plans_dir.is_dir() or not truth_dir.is_dir():
            continue

        # Choose the largest plan PDF under MAX_PLAN_MB (skip tiny RCP excerpts;
        # skip anything bigger than the small-file size band).
        plan_pdf = _pick_plan_pdf(plans_dir)
        # Choose the takeoff xlsx (not Schedule_*).
        truth_xlsx = _pick_truth_xlsx(truth_dir)
        if not plan_pdf or not truth_xlsx:
            continue
        projects.append({
            'name': d.name,
            'plan_pdf': plan_pdf,
            'truth_xlsx': truth_xlsx,
        })
    return projects


def _pick_plan_pdf(plans_dir):
    candidates = list(plans_dir.glob('*.pdf'))
    if not candidates:
        return None
    sized = [(p.stat().st_size, p) for p in candidates
             if p.stat().st_size <= MAX_PLAN_MB * 1024 * 1024]
    if not sized:
        return None
    # Pick the largest under-20MB PDF that doesn't look like specs/RCP excerpt.
    sized.sort(key=lambda t: -t[0])
    for _, p in sized:
        low = p.name.lower()
        if '(specs)' in low or '(rcp)' in low or 'geotech' in low or 'milestones' in low:
            continue
        return p
    return sized[0][1]


def _pick_truth_xlsx(truth_dir):
    candidates = []
    for x in truth_dir.glob('*.xlsx'):
        n = x.name.lower()
        if n.startswith('schedule_') or n.startswith('schdule_'):
            continue
        candidates.append(x)
    if not candidates:
        return None
    # Prefer Takeoff_* if multiple
    for x in candidates:
        if x.name.lower().startswith('takeoff_') or x.name.lower().startswith('takoeff_'):
            return x
    return candidates[0]


# ─── xlsx parsing ──────────────────────────────────────────────────────────

_NORM_TAG_RE = re.compile(r'[^A-Z0-9]')

def _norm_tag(s):
    if not s:
        return ''
    return _NORM_TAG_RE.sub('', str(s).upper())


try:
    from class_aliases import normalize_class
except Exception:
    def normalize_class(s): return s


def _norm_product(s):
    if not s:
        return ''
    name = ' '.join(str(s).upper().split())
    return normalize_class(name)


def read_takeoff_xlsx(path):
    """Read either format. Returns list of {product, tag, qty} dicts.

    Looks for a flat sheet (DATA, RawData) with PRODUCT/QTY/TAG columns. Falls
    back to the first sheet, filtering rows that look like subtotals.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        return [], f'open_failed: {e}'

    # Prefer the TAGGED takeoff sheet. Our output's 'Triune Takeoff' and the
    # team's 'TAKEOFF' carry per-tag rows; 'RawData'/'DATA' is the flat untagged
    # YOLO dump — reading it first discarded all our tags and forced tag_recall=0.
    sheet = None
    for pref in ('Triune Takeoff', 'TAKEOFF', 'TO', 'DATA', 'RawData'):
        if pref in wb.sheetnames:
            sheet = wb[pref]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    # Scan top 10 rows for header. Header row contains "PRODUCT" + a count
    # column (either "QTY" or "COUNT" — the team uses both).
    header_row_idx = None
    headers = None
    for ri, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        cells = [str(c).strip().upper() if c is not None else '' for c in row]
        if 'PRODUCT' in cells and ('QTY' in cells or 'COUNT' in cells):
            header_row_idx = ri
            headers = cells
            break
    if header_row_idx is None:
        return [], 'no_header'

    # Map column indices
    def col_idx(*names):
        for name in names:
            try:
                return headers.index(name)
            except ValueError:
                continue
        return None
    i_prod = col_idx('PRODUCT')
    i_qty  = col_idx('QTY', 'COUNT')
    i_tag  = col_idx('TAG')

    rows = []
    last_product = ''
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or all(c in (None, '') for c in row):
            continue
        prod_raw = row[i_prod] if i_prod is not None and i_prod < len(row) else None
        qty_raw  = row[i_qty]  if i_qty  is not None and i_qty  < len(row) else None
        tag_raw  = row[i_tag]  if i_tag  is not None and i_tag  < len(row) else None

        prod = _norm_product(prod_raw)
        # Carry-down product when team blanks repeated cells
        if prod:
            if 'TOTAL' in prod or prod == 'GRAND TOTAL':
                continue
            last_product = prod
        else:
            prod = last_product
        if not prod:
            continue

        # Parse qty
        try:
            qty = int(qty_raw) if qty_raw not in (None, '') else 0
        except (TypeError, ValueError):
            try:
                qty = int(float(qty_raw))
            except (TypeError, ValueError):
                continue
        if qty <= 0:
            continue
        tag = _norm_tag(tag_raw)
        rows.append({'product': prod, 'tag': tag, 'qty': qty})
    return rows, None


def aggregate_by_product(rows):
    out = defaultdict(int)
    for r in rows:
        out[r['product']] += r['qty']
    return dict(out)


def aggregate_by_product_tag(rows):
    out = defaultdict(int)
    for r in rows:
        out[(r['product'], r['tag'])] += r['qty']
    return dict(out)


def aggregate_by_tag(rows):
    """Aggregate by TAG only (ignore product label). The team and our pipeline
    label the same equipment differently (e.g. 'FAN' vs 'EXHAUST FAN'), so the
    (product, tag) tuple zeroes matches even when tags are identical. This
    tag-only view measures whether the tag itself was found."""
    out = defaultdict(int)
    for r in rows:
        if r['tag']:
            out[r['tag']] += r['qty']
    return dict(out)


# ─── Scoring ───────────────────────────────────────────────────────────────

def _overlap_score(team, ours):
    """Generic overlap of two qty dicts. Returns recall, precision, match_total."""
    keys = set(team) | set(ours)
    match = 0
    team_total = sum(team.values())
    ours_total = sum(ours.values())
    for k in keys:
        match += min(team.get(k, 0), ours.get(k, 0))
    recall = match / team_total if team_total else 0.0
    precision = match / ours_total if ours_total else 0.0
    return recall, precision, match


def score_project(team_rows, our_rows):
    team_p = aggregate_by_product(team_rows)
    our_p  = aggregate_by_product(our_rows)
    p_recall, p_prec, p_match = _overlap_score(team_p, our_p)

    team_pt = aggregate_by_product_tag(team_rows)
    our_pt  = aggregate_by_product_tag(our_rows)
    t_recall, t_prec, t_match = _overlap_score(team_pt, our_pt)

    # Tag-only (product-label-independent) — the honest measure of tag matching.
    team_t = aggregate_by_tag(team_rows)
    our_t  = aggregate_by_tag(our_rows)
    to_recall, to_prec, to_match = _overlap_score(team_t, our_t)

    return {
        'team_total': sum(team_p.values()),
        'our_total':  sum(our_p.values()),
        'product_recall':    p_recall,
        'product_precision': p_prec,
        'product_match':     p_match,
        'tag_recall':    t_recall,        # strict: (product, tag) tuple
        'tag_precision': t_prec,
        'tag_match':     t_match,
        'tag_only_recall':    to_recall,  # lenient: tag string only
        'tag_only_precision': to_prec,
        'tag_only_match':     to_match,
        'team_products': team_p,
        'our_products':  our_p,
    }


# ─── Pipeline runner ───────────────────────────────────────────────────────

def run_one_project(project, out_root, cache=False, model=None):
    """Run takeoff_cli.py on the project's plan PDF. Returns dict with status,
    runtime, our_xlsx_path, error."""
    name = project['name']
    pdf = project['plan_pdf']
    out_dir = out_root / _safe_dir(name)
    out_dir.mkdir(parents=True, exist_ok=True)
    our_xlsx = out_dir / f"{pdf.stem}_takeoff.xlsx"
    variables_json = out_dir / f"{pdf.stem}_variables.json"

    if cache and our_xlsx.exists():
        return {'status': 'cached', 'our_xlsx': our_xlsx, 'variables_json': variables_json,
                'runtime_s': 0.0, 'error': None}

    cmd = [
        sys.executable, str(TOOL_DIR / 'takeoff_cli.py'),
        str(pdf), '--output-dir', str(out_dir),
    ]
    if model:
        cmd += ['--model', str(model)]
    t0 = time.time()
    try:
        # 10-min timeout per project. UTF-8 for non-ASCII project names.
        proc = subprocess.run(cmd, capture_output=True, timeout=600,
                              encoding='utf-8', errors='replace')
    except subprocess.TimeoutExpired:
        return {'status': 'timeout', 'our_xlsx': our_xlsx, 'variables_json': variables_json,
                'runtime_s': time.time() - t0, 'error': 'CLI timed out (>10min)'}
    runtime = time.time() - t0

    if proc.returncode != 0:
        tail = (proc.stderr or '')[-400:].replace('\n', ' | ')
        return {'status': 'crashed', 'our_xlsx': our_xlsx, 'variables_json': variables_json,
                'runtime_s': runtime, 'error': tail or f'exit={proc.returncode}'}

    if not our_xlsx.exists():
        # CLI may exit 0 with no detections (e.g. blank pages). Check stdout.
        if 'No HVAC equipment detected' in (proc.stdout or ''):
            return {'status': 'no_detections', 'our_xlsx': None,
                    'variables_json': variables_json,
                    'runtime_s': runtime, 'error': None}
        return {'status': 'no_xlsx', 'our_xlsx': None,
                'variables_json': variables_json, 'runtime_s': runtime,
                'error': 'CLI completed but no xlsx written'}

    return {'status': 'ok', 'our_xlsx': our_xlsx, 'variables_json': variables_json,
            'runtime_s': runtime, 'error': None}


_DIR_SAFE_RE = re.compile(r'[^A-Za-z0-9._\- ]+')

def _safe_dir(name):
    return _DIR_SAFE_RE.sub('_', name).strip()[:120]


def schedule_tag_count(variables_json):
    if not variables_json or not variables_json.exists():
        return 0
    try:
        with open(variables_json, encoding='utf-8') as f:
            data = json.load(f)
        return len({v.get('tag') for v in data if v.get('tag')})
    except Exception:
        return 0


# ─── Reports ───────────────────────────────────────────────────────────────

RESULTS_HEADERS = [
    'project', 'status', 'team_total', 'our_total', 'our_tagged_pct_of_team',
    'product_recall', 'product_precision',
    'tag_recall', 'tag_precision',
    'schedule_tags_found', 'runtime_s', 'error',
]

PER_PRODUCT_HEADERS = [
    'project', 'product', 'team_qty', 'our_qty', 'match_qty', 'over', 'under',
]


def write_results_csv(rows, path):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=RESULTS_HEADERS)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, '') for k in RESULTS_HEADERS})


def write_per_product_csv(rows, path):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=PER_PRODUCT_HEADERS)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_summary_md(results, path):
    scored = [r for r in results if r.get('product_recall') is not None]
    scored.sort(key=lambda r: -r['product_recall'])

    n_total = len(results)
    n_scored = len(scored)
    n_above_50 = sum(1 for r in scored if r['product_recall'] >= 0.5)
    n_above_25 = sum(1 for r in scored if r['product_recall'] >= 0.25)
    n_zero = sum(1 for r in scored if r['product_recall'] == 0)

    if scored:
        recalls = sorted(r['product_recall'] for r in scored)
        median = recalls[len(recalls) // 2]
        mx = max(recalls)
    else:
        median = mx = 0.0

    fail_modes = defaultdict(int)
    for r in results:
        fail_modes[r['status']] += 1

    lines = []
    lines.append(f"# Benchmark Summary")
    lines.append("")
    lines.append(f"- Total projects discovered: **{n_total}**")
    lines.append(f"- Successfully scored:       **{n_scored}**")
    lines.append(f"- Median product recall:     **{median:.0%}**")
    lines.append(f"- Max product recall:        **{mx:.0%}**")
    lines.append(f"- Projects ≥ 50% recall:     **{n_above_50}**")
    lines.append(f"- Projects ≥ 25% recall:     **{n_above_25}**")
    lines.append(f"- Projects at 0%:            **{n_zero}**")
    lines.append("")
    lines.append("## Status breakdown")
    lines.append("")
    lines.append("| status | count |")
    lines.append("|---|---|")
    for k, v in sorted(fail_modes.items(), key=lambda x: -x[1]):
        lines.append(f"| {k} | {v} |")
    lines.append("")

    def _table(rows):
        lines.append("| project | recall | precision | team | ours | tagged% | sched_tags | status |")
        lines.append("|---|---:|---:|---:|---:|---:|---:|---|")
        for r in rows:
            lines.append(
                f"| {r['project'][:50]} "
                f"| {r['product_recall']:.0%} "
                f"| {r['product_precision']:.0%} "
                f"| {r['team_total']} "
                f"| {r['our_total']} "
                f"| {r['our_tagged_pct_of_team']:.0%} "
                f"| {r['schedule_tags_found']} "
                f"| {r['status']} |"
            )

    lines.append(f"## Top 10 by product recall")
    lines.append("")
    _table(scored[:10])
    lines.append("")
    lines.append(f"## Bottom 10 by product recall")
    lines.append("")
    _table(scored[-10:][::-1])
    lines.append("")

    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


# ─── Main ──────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--root', default=str(SAMPLE_ROOT))
    ap.add_argument('--projects', nargs='*', default=None,
                    help='Substring filter — only projects whose name contains any')
    ap.add_argument('--cache', action='store_true',
                    help='Skip projects with existing output xlsx')
    ap.add_argument('--limit', type=int, default=None,
                    help='Process at most N projects (debugging)')
    ap.add_argument('--model', default=None,
                    help='Path to YOLO model .pt (passes through to takeoff_cli --model). '
                         'Default: takeoff_cli uses models/hvac_yolov8s_v9.pt.')
    args = ap.parse_args()

    root = Path(args.root)
    OUT_ROOT.mkdir(parents=True, exist_ok=True)

    projects = discover_projects(root)
    if args.projects:
        wanted = [w.lower() for w in args.projects]
        projects = [p for p in projects if any(w in p['name'].lower() for w in wanted)]
    if args.limit:
        projects = projects[:args.limit]

    print(f"Discovered {len(projects)} project(s) to benchmark")
    print(f"Output: {OUT_ROOT}")
    print()

    results = []
    per_product = []
    t_run = time.time()

    for i, proj in enumerate(projects, 1):
        name = proj['name']
        print(f"[{i}/{len(projects)}] {name}")
        print(f"  plan:  {proj['plan_pdf'].name}")
        print(f"  truth: {proj['truth_xlsx'].name}")

        run_info = run_one_project(proj, OUT_ROOT, cache=args.cache, model=args.model)
        status = run_info['status']
        print(f"  status: {status}  ({run_info['runtime_s']:.0f}s)")

        # Always read team xlsx
        team_rows, team_err = read_takeoff_xlsx(proj['truth_xlsx'])
        team_total = sum(r['qty'] for r in team_rows)

        # Read our xlsx if pipeline produced one
        our_rows, our_err = [], None
        if status in ('ok', 'cached') and run_info['our_xlsx']:
            our_rows, our_err = read_takeoff_xlsx(run_info['our_xlsx'])
        our_total = sum(r['qty'] for r in our_rows)

        sched_tags = schedule_tag_count(run_info.get('variables_json'))

        if team_rows and our_rows:
            score = score_project(team_rows, our_rows)
            row = {
                'project': name,
                'status': status,
                'team_total': team_total,
                'our_total':  our_total,
                'our_tagged_pct_of_team': (
                    score['product_match'] / team_total if team_total else 0.0
                ),
                'product_recall':    score['product_recall'],
                'product_precision': score['product_precision'],
                'tag_recall':    score['tag_recall'],
                'tag_precision': score['tag_precision'],
                'schedule_tags_found': sched_tags,
                'runtime_s': round(run_info['runtime_s'], 1),
                'error': run_info['error'] or '',
            }
            for prod in set(score['team_products']) | set(score['our_products']):
                tq = score['team_products'].get(prod, 0)
                oq = score['our_products'].get(prod, 0)
                per_product.append({
                    'project': name,
                    'product': prod,
                    'team_qty': tq,
                    'our_qty':  oq,
                    'match_qty': min(tq, oq),
                    'over':  max(0, oq - tq),
                    'under': max(0, tq - oq),
                })
            print(f"  recall: {score['product_recall']:.0%}  "
                  f"precision: {score['product_precision']:.0%}  "
                  f"team={team_total} ours={our_total}")
        else:
            row = {
                'project': name,
                'status': status,
                'team_total': team_total,
                'our_total':  our_total,
                'our_tagged_pct_of_team': 0.0,
                'product_recall':    None,
                'product_precision': None,
                'tag_recall':    None,
                'tag_precision': None,
                'schedule_tags_found': sched_tags,
                'runtime_s': round(run_info['runtime_s'], 1),
                'error': run_info['error'] or team_err or our_err or '',
            }
            print(f"  unscored — team={team_total} ours={our_total} err={row['error']}")

        results.append(row)

        # Incremental write so a long run is recoverable
        write_results_csv(results, OUT_ROOT / 'benchmark_results.csv')
        write_per_product_csv(per_product, OUT_ROOT / 'benchmark_per_product.csv')
        print()

    write_summary_md(results, OUT_ROOT / 'benchmark_summary.md')
    elapsed = time.time() - t_run
    print(f"\nDone in {elapsed/60:.1f} min")
    print(f"  benchmark_results.csv      {OUT_ROOT/'benchmark_results.csv'}")
    print(f"  benchmark_per_product.csv  {OUT_ROOT/'benchmark_per_product.csv'}")
    print(f"  benchmark_summary.md       {OUT_ROOT/'benchmark_summary.md'}")


if __name__ == '__main__':
    main()
