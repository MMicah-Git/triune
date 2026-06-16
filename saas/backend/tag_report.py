"""
tag_report.py — tag-by-tag breakdown report.

Combines three signals into one table:
  1. Per-instance plan-label extraction (diffuser_extractor.py — reads
     individual GRD mark+neck+CFM from plan text labels). NEW: pulled in
     from the sibling hvac-takeoff-tool checkout.
  2. YOLO detection counts (grouped by AI class + tag from detections.json).
  3. Schedule cross-reference (when variables.json has tag data).

Produces:
  - <stem>_tag_report.json    (structured, for programmatic use)
  - <stem>_tag_report.md      (Markdown table — renders inline in the UI)
  - <stem>_tag_report.xlsx    (Excel — same shape as the team's takeoff)

Each row in the report represents ONE tag, with:
  • Source        — where the tag came from (plan-label / yolo / schedule)
  • Tag           — A1, CU-1, RG-2, etc.
  • Class         — AI class or schedule's inferred class
  • Detected qty  — how many YOLO detections were assigned this tag
  • Plan qty      — how many plan-label instances we found
  • CFM (total)   — summed from plan labels
  • Neck size     — from plan label or schedule
  • Manufacturer / Model / Module size  — from schedule
  • Discrepancy   — flag if Detected ≠ Plan or schedule mismatch
"""

from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

HERE = Path(__file__).resolve().parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))


def _safe_call_diffuser_extractor(pdf_path: Path, plan_pages_1based: list[int]):
    """Call diffuser_extractor.extract_diffuser_instances safely.
    Returns ([], []) if it fails or finds nothing.
    Returns (instances, warnings) on success."""
    try:
        from diffuser_extractor import extract_diffuser_instances
        plan_pages_0based = [p - 1 for p in plan_pages_1based]
        return extract_diffuser_instances(pdf_path, plan_pages_0based)
    except Exception as e:
        return [], [f'diffuser_extractor failed: {e}']


def _safe_call_bom(instances, mark_details=None):
    try:
        from diffuser_extractor import aggregate_diffuser_bom
        return aggregate_diffuser_bom(instances, mark_details=mark_details)
    except Exception:
        return []


def _normalize_neck(s: str | None) -> str:
    """Normalize neck-size formats so they compare equal across sources.

    Examples:
      'round:8' / '8\"' / '8 INCH' → '8'
      'rect:10x10' / '10X10' / '10/10' → '10x10'
      '' / None → ''
    """
    if not s:
        return ''
    s = str(s).strip()
    if s.lower().startswith('round:'):
        s = s[6:]
    elif s.lower().startswith('rect:'):
        s = s[5:]
    # Strip quotes/units
    s = s.replace('"', '').replace("'", '').replace('″', '').strip()
    s = s.replace(' INCH', '').replace(' inch', '').strip()
    # Standardize rect separator
    import re as _re
    m = _re.match(r'^(\d+(?:\.\d+)?)\s*[xX/]\s*(\d+(?:\.\d+)?)$', s)
    if m:
        return f'{m.group(1)}x{m.group(2)}'.lower()
    return s


def _neck_size_from_schedule(props: dict) -> str:
    """Pull neck size from a schedule's properties dict. Tries several
    common header variants: NECK SIZE, NECK, INLET SIZE, etc."""
    if not props:
        return ''
    for k, v in props.items():
        ku = (k or '').upper()
        # Match canonical neck-size column names
        if ku in ('NECK SIZE', 'NECK', 'INLET SIZE', 'INLET',
                  'SIZE (NECK)', 'CONNECTION SIZE', 'CONN SIZE'):
            return str(v or '').strip()
    # Fallback: any key that ends with " NECK" or starts with "NECK"
    for k, v in props.items():
        ku = (k or '').upper().strip()
        if ku.startswith('NECK') and 'TAB' not in ku:
            return str(v or '').strip()
    return ''


def _effective_count(d: dict) -> int:
    """Billable units for one detection after TYP/NIC plan-note semantics.

      • NIC ("not in contract" / "by others") → 0 (carved out of the bid)
      • "(TYP OF N)" explicit multiplier        → N
      • everything else                          → 1
    Marks are set upstream by typ_uno_nic.apply_typ_uno_nic.
    """
    if d.get('nic'):
        return 0
    m = d.get('typ_of_count')
    if isinstance(m, int) and m > 1:
        return m
    return 1


def build_tag_table(
    pdf_path: Path,
    detections: dict,
    variables: list[dict],
    plan_pages_1based: list[int],
    use_ocr_fallback: bool = True,
) -> dict:
    """Build the unified tag-by-tag breakdown."""

    # ── Source 1a: per-instance plan-label extraction (text layer) ──────────
    instances, plan_warnings = _safe_call_diffuser_extractor(pdf_path, plan_pages_1based)

    # ── Source 1b: OCR fallback for raster plans ────────────────────────────
    # Only fire if text-layer found nothing AND we have YOLO detections.
    n_detections = sum(len(v) for v in detections.get('pages', {}).values())
    if use_ocr_fallback and not instances and n_detections > 0:
        try:
            from plan_label_ocr import extract_plan_labels_via_ocr, merge_with_diffuser_extractor
            print(f'[tag_report] text-layer plan-label extraction found 0; running OCR fallback over {n_detections} detections')
            ocr_instances = extract_plan_labels_via_ocr(
                pdf_path, detections, plan_pages_1based, dpi=300,
            )
            instances = merge_with_diffuser_extractor(instances, ocr_instances)
            if ocr_instances:
                plan_warnings.append(f'OCR fallback recovered {len(ocr_instances)} plan-label candidate(s)')
        except Exception as e:
            plan_warnings.append(f'OCR fallback failed: {e}')

    # Build mark_details from variables (schedule data) for the join
    mark_details = {}
    for v in variables or []:
        tag = (v.get('tag') or '').upper()
        if tag:
            mark_details[tag] = v.get('properties') or {}

    bom = _safe_call_bom(instances, mark_details=mark_details)

    # ── Source 2: YOLO detection counts grouped by (class, tag) ─────────────
    # Also collect waterfall neck-size + confidence per (class, tag, neck) bucket.
    yolo_class_tag_counts: dict[tuple[str, str | None], int] = Counter()
    waterfall_by_class_tag_neck: dict[tuple[str, str, str], dict] = {}
    nic_excluded = 0          # detections carved out by NIC / by others
    typ_of_extra = 0          # extra units added by "(TYP OF N)" multipliers
    for pkey, det_list in detections.get('pages', {}).items():
        for d in det_list:
            cls = d.get('cls', '?')
            tag = d.get('tag')
            eff = _effective_count(d)
            if eff == 0:
                nic_excluded += 1
                continue       # NIC: out of all counts (still listed in typ_uno_nic.json)
            typ_of_extra += eff - 1
            yolo_class_tag_counts[(cls, tag)] += eff

            # Waterfall enrichment fields (set by neck_size_waterfall_runner).
            # Buckets stay at physical count (+1); any TYP-OF extra units fall
            # through to the residual yolo rows below, keeping the math consistent.
            neck = d.get('neck_size') or ''
            tier = d.get('neck_tier') or ''
            conf = float(d.get('neck_confidence') or 0)
            source = d.get('neck_source') or ''

            if neck:
                key = (cls, (tag or '').upper(), neck)
                bucket = waterfall_by_class_tag_neck.setdefault(key, {
                    'count': 0, 'max_conf': 0.0, 'min_conf': 1.0,
                    'tiers': Counter(), 'sources': Counter(),
                })
                bucket['count'] += 1
                bucket['max_conf'] = max(bucket['max_conf'], conf)
                bucket['min_conf'] = min(bucket['min_conf'], conf)
                if tier:
                    bucket['tiers'][tier] += 1
                if source:
                    bucket['sources'][source] += 1

    # ── Source 3: schedule tags (variables) ─────────────────────────────────
    schedule_tags = {(v.get('tag') or '').upper() for v in (variables or []) if v.get('tag')}

    # ── Combine into a single tag-by-tag table ──────────────────────────────
    rows: list[dict] = []

    # Rows from BOM (plan-label extractor)
    seen_keys = set()
    for b in bom:
        tag = (b.get('mark') or '').upper()
        key = (tag, b.get('neck_size_canon') or '')
        seen_keys.add(key)
        # Find AI detections matching this tag
        yolo_count = sum(n for (_cls, t), n in yolo_class_tag_counts.items()
                         if (t or '').upper() == tag)
        discrepancy = []
        if yolo_count and b['qty'] != yolo_count:
            discrepancy.append(f'plan={b["qty"]} vs yolo={yolo_count}')
        if b.get('cfm_missing', 0) > 0:
            discrepancy.append(f'{b["cfm_missing"]} instance(s) missing CFM')
        if tag and tag not in schedule_tags and variables:
            discrepancy.append('tag not in schedule')
        # Neck size: prefer plan-label (per-instance, like "round:8") and
        # fall back to schedule's NECK SIZE column when plan didn't say.
        neck = b.get('neck_size_canon') or _neck_size_from_schedule(mark_details.get(tag, {}))
        neck_norm = _normalize_neck(neck)

        # Find waterfall buckets for this (tag, normalized-neck) — bring
        # confidence + tier + per-(tag,neck) yolo count into the BOM row.
        # Sum across classes since the BOM row represents the tag as a whole.
        wf_tier = ''
        wf_conf = 0.0
        wf_yolo_per_neck = 0
        wf_source_detail = ''
        for (w_cls, w_tag, w_neck), wb in waterfall_by_class_tag_neck.items():
            if w_tag == tag and _normalize_neck(w_neck) == neck_norm:
                wf_yolo_per_neck += wb['count']
                if wb['max_conf'] > wf_conf:
                    wf_conf = wb['max_conf']
                    wf_tier = (wb['tiers'].most_common(1)[0][0]
                              if wb['tiers'] else '')
                    wf_source_detail = (wb['sources'].most_common(1)[0][0]
                                       if wb['sources'] else '')

        # Display the cleaner normalized form
        display_neck = neck_norm or neck or ''
        if display_neck and not display_neck.endswith('"') and 'x' not in display_neck:
            display_neck = display_neck + '"'

        # Prefer per-(tag, neck) yolo count when waterfall provides it;
        # falls back to the per-tag aggregate when we have nothing better.
        per_neck_yolo = wf_yolo_per_neck if wf_yolo_per_neck > 0 else yolo_count

        rows.append({
            'tag': tag,
            'class': '',  # filled below if we can correlate
            'neck_size': display_neck,
            'neck_norm': neck_norm,  # for dedup later
            'neck_tier': wf_tier,
            'neck_confidence': round(wf_conf, 2) if wf_conf else 0,
            'plan_qty': b['qty'],
            'yolo_qty': per_neck_yolo,
            'total_cfm': b.get('total_cfm', 0),
            'cfm_missing': b.get('cfm_missing', 0),
            'manufacturer': b.get('manufacturer', ''),
            'model': b.get('model', ''),
            'module_size': b.get('module_size', ''),
            'mounting': b.get('mounting', ''),
            'source': 'plan+schedule' if b.get('manufacturer') else 'plan',
            'discrepancy': '; '.join(discrepancy) if discrepancy else '',
        })

    # Waterfall-grouped rows — one row per (class, tag, neck_size) where
    # the waterfall succeeded. Skips tags already covered by plan-label BOM.
    waterfall_tags_emitted = set()
    for (cls, tag_u, neck), bucket in sorted(waterfall_by_class_tag_neck.items()):
        neck_norm = _normalize_neck(neck)
        # Skip if a BOM row already covered this (tag, normalized-neck)
        if any(r['tag'] == (tag_u or '') and r.get('neck_norm') == neck_norm
               for r in rows if r.get('neck_norm') is not None):
            continue
        if (tag_u, neck_norm) in waterfall_tags_emitted:
            continue
        waterfall_tags_emitted.add((tag_u, neck_norm))

        # Confidence summary for the bucket
        max_conf = bucket['max_conf']
        # Determine row tier from majority
        tier = bucket['tiers'].most_common(1)[0][0] if bucket['tiers'] else ''
        # Source: pick the most-used source
        src_short = bucket['sources'].most_common(1)[0][0] if bucket['sources'] else ''
        # Shorten source for display
        src_short = src_short.replace('level1-plan-text-combined-', 'plan-')\
                              .replace('level2-schedule-via-bubble', 'schedule')\
                              .replace('level3-ocr-near-detection', 'ocr')\
                              .replace('level4-cfm-range-lookup', 'cfm-lookup')

        row = {
            'tag': tag_u or '',
            'class': cls,
            'neck_size': neck,
            'plan_qty': 0,
            'yolo_qty': bucket['count'],
            'total_cfm': 0,
            'cfm_missing': 0,
            'manufacturer': '',
            'model': '',
            'module_size': '',
            'mounting': '',
            'source': 'waterfall' if not tag_u else f'waterfall+{src_short}',
            'discrepancy': '',
            'neck_tier': tier,
            'neck_confidence': round(max_conf, 2),
            'neck_source_detail': src_short,
        }
        # Pull schedule props if tag matches
        if tag_u and tag_u in mark_details:
            props = mark_details[tag_u]
            row['manufacturer'] = props.get('MANUFACTURER', '') or props.get('MAKE', '')
            row['model'] = props.get('MODEL', '')
            row['module_size'] = props.get('MODULE SIZE', '')
            row['mounting'] = props.get('MOUNTING', '')
        rows.append(row)

    # Untagged / no-neck residual: detections that didn't get tag OR neck.
    # Emit them as one row per class with the residual count.
    yolo_only_tags: dict[str, Counter] = defaultdict(Counter)
    for (cls, tag), n in yolo_class_tag_counts.items():
        tag_u = (tag or '').upper() if tag else None
        # Subtract waterfall-emitted counts for this (cls, tag)
        waterfall_counted = sum(
            b['count'] for (c, t, _), b in waterfall_by_class_tag_neck.items()
            if c == cls and t == (tag_u or '')
        )
        remaining = n - waterfall_counted
        if remaining > 0:
            yolo_only_tags[tag_u][cls] += remaining

    for tag_u, class_counts in yolo_only_tags.items():
        if not class_counts:
            continue
        already = any(r['tag'] == (tag_u or '') and not r.get('neck_size') for r in rows)
        if already and tag_u:
            continue
        for cls, n in class_counts.items():
            row = {
                'tag': tag_u or '',
                'class': cls,
                'neck_size': '',
                'plan_qty': 0,
                'yolo_qty': n,
                'total_cfm': 0,
                'cfm_missing': 0,
                'manufacturer': '',
                'model': '',
                'module_size': '',
                'mounting': '',
                'source': 'yolo' if tag_u else 'yolo-untagged',
                'discrepancy': 'no neck size found' if tag_u else 'untagged',
                'neck_tier': 'LOW',
                'neck_confidence': 0.0,
            }
            if tag_u and tag_u in mark_details:
                props = mark_details[tag_u]
                row['manufacturer'] = props.get('MANUFACTURER', '') or props.get('MAKE', '')
                row['model'] = props.get('MODEL', '')
                row['module_size'] = props.get('MODULE SIZE', '')
                row['mounting'] = props.get('MOUNTING', '')
                row['neck_size'] = _neck_size_from_schedule(props)
                row['source'] = 'yolo+schedule'
            rows.append(row)

    # Rows for schedule tags with NO matching detection
    for tag_u in schedule_tags:
        if any(r['tag'] == tag_u for r in rows):
            continue
        props = mark_details.get(tag_u, {})
        rows.append({
            'tag': tag_u,
            'class': '',
            'neck_size': _neck_size_from_schedule(props),
            'plan_qty': 0,
            'yolo_qty': 0,
            'total_cfm': 0,
            'cfm_missing': 0,
            'manufacturer': props.get('MANUFACTURER', '') or props.get('MAKE', ''),
            'model': props.get('MODEL', ''),
            'module_size': props.get('MODULE SIZE', ''),
            'mounting': props.get('MOUNTING', ''),
            'source': 'schedule-only',
            'discrepancy': 'scheduled but no detection',
        })

    # Sort: rows with discrepancies first, then by tag
    rows.sort(key=lambda r: (not r['discrepancy'], r['tag'] or 'zzz'))

    return {
        'rows': rows,
        'totals': {
            'rows': len(rows),
            'tagged_detections': sum(r['yolo_qty'] for r in rows if r['tag']),
            'untagged_detections': sum(r['yolo_qty'] for r in rows if not r['tag']),
            'plan_instances': sum(r['plan_qty'] for r in rows),
            'total_cfm': sum(r['total_cfm'] for r in rows),
            'rows_with_discrepancy': sum(1 for r in rows if r['discrepancy']),
            'nic_excluded': nic_excluded,
            'typ_of_extra_units': typ_of_extra,
        },
        'plan_label_warnings': plan_warnings,
    }


def render_markdown_table(table: dict) -> str:
    """Render the tag-by-tag table as Markdown."""
    rows = table['rows']
    totals = table['totals']
    if not rows:
        return '_No tag data to report. The plan-label extractor found nothing readable and no YOLO detections were tagged._'

    lines = []
    lines.append(f'**{totals["rows"]} tag rows**  ·  '
                 f'{totals["tagged_detections"]} tagged + '
                 f'{totals["untagged_detections"]} untagged YOLO detections  ·  '
                 f'{totals["plan_instances"]} plan-label instances  ·  '
                 f'{totals["total_cfm"]} total CFM  ·  '
                 f'**{totals["rows_with_discrepancy"]} discrepancies**')
    if totals.get('nic_excluded') or totals.get('typ_of_extra_units'):
        lines.append('')
        lines.append(f'_Plan-note adjustments: '
                     f'{totals.get("nic_excluded", 0)} NIC detection(s) excluded · '
                     f'+{totals.get("typ_of_extra_units", 0)} unit(s) added from "(TYP OF N)" markers._')
    lines.append('')
    lines.append('| Tag | Class | Neck | Conf | YOLO qty | Plan qty | Mfr | Model | Module | Mounting | Source |')
    lines.append('|-----|-------|------|------|---------:|---------:|-----|-------|--------|----------|--------|')

    tier_emoji = {'HIGH': '🟢', 'MEDIUM': '🟡', 'LOW': '🔴'}

    for r in rows[:120]:  # cap at 120 rows in markdown
        tier = r.get('neck_tier', '')
        emoji = tier_emoji.get(tier, '')
        conf_val = r.get('neck_confidence', 0)
        conf_display = f"{emoji} {conf_val:.2f}" if conf_val else '—'

        cells = [
            r['tag'] or '—',
            r['class'] or '—',
            r['neck_size'] or '—',
            conf_display,
            str(r['yolo_qty']),
            str(r['plan_qty']),
            r['manufacturer'] or '—',
            r['model'] or '—',
            r['module_size'] or '—',
            r['mounting'] or '—',
            r['source'],
        ]
        lines.append('| ' + ' | '.join(str(c)[:60] for c in cells) + ' |')
    if len(rows) > 120:
        lines.append(f'_… {len(rows) - 120} more rows truncated. Download the JSON or Excel for the full table._')
    return '\n'.join(lines)


def render_excel(table: dict, output_path: Path) -> None:
    """Write the tag-by-tag table as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tag-by-Tag'

    headers = ['Tag', 'Class', 'Neck Size', 'Conf', 'Tier',
               'YOLO Qty', 'Plan Qty', 'Total CFM',
               'CFM Missing', 'Manufacturer', 'Model', 'Module Size', 'Mounting',
               'Source', 'Discrepancy']
    keys = ['tag', 'class', 'neck_size', 'neck_confidence', 'neck_tier',
            'yolo_qty', 'plan_qty', 'total_cfm',
            'cfm_missing', 'manufacturer', 'model', 'module_size', 'mounting',
            'source', 'discrepancy']

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    discrepancy_fill = PatternFill('solid', fgColor='FFD966')

    # 3-tier neck-confidence colors (matches PLAN.md §5)
    high_fill = PatternFill('solid', fgColor='C6EFCE')     # green
    medium_fill = PatternFill('solid', fgColor='FFEB9C')   # yellow
    low_fill = PatternFill('solid', fgColor='FFC7CE')      # red/pink

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for ri, r in enumerate(table['rows'], 2):
        tier = r.get('neck_tier', '')
        for ci, key in enumerate(keys, 1):
            cell = ws.cell(row=ri, column=ci, value=r.get(key, ''))
            # Color the neck-size + tier cells by tier
            if key in ('neck_size', 'neck_confidence', 'neck_tier'):
                if tier == 'HIGH':
                    cell.fill = high_fill
                elif tier == 'MEDIUM':
                    cell.fill = medium_fill
                elif tier == 'LOW':
                    cell.fill = low_fill
            elif r.get('discrepancy'):
                cell.fill = discrepancy_fill

    # Auto-width
    for ci, h in enumerate(headers, 1):
        max_w = max(len(h),
                   *[len(str(r.get(keys[ci-1], ''))) for r in table['rows']])
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_w + 2, 40)

    # Summary sheet
    ws2 = wb.create_sheet('Summary')
    for ri, (k, v) in enumerate(table['totals'].items(), 1):
        ws2.cell(row=ri, column=1, value=k.replace('_', ' ').title()).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=v)

    wb.save(str(output_path))


def write_tag_report(
    pdf_path: Path,
    detections: dict,
    variables: list[dict],
    plan_pages_1based: list[int],
    output_dir: Path,
    stem: str,
) -> dict:
    """Top-level entry: build the table, write JSON + Markdown + Excel."""
    table = build_tag_table(pdf_path, detections, variables, plan_pages_1based)
    (output_dir / f'{stem}_tag_report.json').write_text(
        json.dumps(table, indent=2, default=str), encoding='utf-8')
    md = render_markdown_table(table)
    (output_dir / f'{stem}_tag_report.md').write_text(
        f'# Tag-by-Tag Report\n\n{md}\n', encoding='utf-8')
    try:
        render_excel(table, output_dir / f'{stem}_tag_report.xlsx')
        xlsx_made = True
    except Exception as e:
        print(f'[tag_report] Excel write failed: {e}')
        xlsx_made = False

    return {
        'totals': table['totals'],
        'artifacts': {
            'tag_report_json': f'{stem}_tag_report.json',
            'tag_report_md': f'{stem}_tag_report.md',
            **({'tag_report_xlsx': f'{stem}_tag_report.xlsx'} if xlsx_made else {}),
        }
    }


if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument('pdf')
    ap.add_argument('--detections', required=True)
    ap.add_argument('--variables')
    ap.add_argument('--plan-pages', nargs='+', type=int, required=True)
    ap.add_argument('--out-dir', required=True)
    args = ap.parse_args()

    dets = json.loads(Path(args.detections).read_text(encoding='utf-8'))
    vars_ = json.loads(Path(args.variables).read_text(encoding='utf-8')) if args.variables else []
    if isinstance(vars_, dict): vars_ = vars_.get('variables', []) or []

    result = write_tag_report(Path(args.pdf), dets, vars_,
                             args.plan_pages,
                             Path(args.out_dir), Path(args.pdf).stem)
    print(json.dumps(result, indent=2))
