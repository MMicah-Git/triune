"""
compare_excel.py — GRD takeoff comparison tool.

Compares the team's completed takeoff Excel against our generated takeoff Excel
and writes a multi-sheet comparison_report.xlsx.

Usage
-----
python compare_excel.py --team <team.xlsx> --ours <ours.xlsx> [--out <report.xlsx>]

Default paths are the Busy Bees project files.
"""

import argparse
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour palette ────────────────────────────────────────────────────────────

_FILL = {
    'match':   PatternFill('solid', fgColor='C6EFCE'),   # green
    'diff':    PatternFill('solid', fgColor='FFEB9C'),   # yellow
    'missing': PatternFill('solid', fgColor='FFC7CE'),   # red/pink
    'extra':   PatternFill('solid', fgColor='BDD7EE'),   # blue
    'header':  PatternFill('solid', fgColor='D9D9D9'),   # grey
    'title':   PatternFill('solid', fgColor='4472C4'),   # dark blue
    'sub':     PatternFill('solid', fgColor='9DC3E6'),   # mid blue
}
_BOLD  = Font(bold=True)
_WHITE = Font(bold=True, color='FFFFFF')
_THIN  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)


# ── Normalisation helpers ─────────────────────────────────────────────────────

def _norm_dim(s) -> str:
    """
    Normalise neck/module dimension strings for comparison.

    Handles all inch-mark variants and dimension separators:
      '10X10'   → '10x10'
      '10×10'   → '10x10'   (Unicode ×)
      '10/10'   → '10x10'
      '24"X24"' → '24x24'
      "12''"    → '12'      (CAD double-apostrophe inch mark)
      '12"'     → '12'      (round — bare diameter)
      '.'       → '.'       (null placeholder — preserved as-is)
    """
    if s is None:
        return ''
    s = str(s).strip()
    if s in ('.', ''):                        # null placeholders — don't alter
        return s
    s = s.replace('"', '').replace("'", '')   # strip all inch marks (" and ')
    s = s.replace('×', 'x')             # Unicode × → x
    s = s.replace('X', 'x')                  # capital X → x
    s = s.replace('/', 'x')                  # slash → x  (10/10 → 10x10)
    return s.strip()


# Compound tag normaliser: strips purely-alphabetic word suffixes from
# hyphenated marks so that team tags like 'LD-1-PLENUM' and 'LD-1-SLOT DIFFUSER'
# compare equal to our extracted base mark 'LD-1'.
#
# Rules:
#   LD-1-PLENUM        → LD-1      (single-word suffix stripped)
#   LD-1-SLOT DIFFUSER → LD-1      (multi-word suffix stripped)
#   VAV-1A             → VAV-1A    (single-letter suffix kept — unique id)
#   SG-1               → SG-1      (no suffix)
#   HP-1A              → HP-1A     (single letter kept)
#
# Only strips when the suffix after the base LETTER-DIGIT mark is composed of
# 2+ uppercase letters (optionally followed by more space-separated words).
_COMPOUND_TAG_RE = re.compile(
    r'^([A-Z]{1,4}-?\d{1,4})'          # base mark: LETTERS-DIGITS
    r'(?:-[A-Z]{2,}(?:\s+[A-Z]{2,})*'  # first word suffix (≥2 letters)
    r'(?:\s+[A-Z]+)*)?$',               # optional additional words
    re.IGNORECASE,
)


def _norm_tag(s) -> str:
    if s is None:
        return ''
    raw = str(s).strip().upper()
    # Strip trailing parenthetical like 'CD-1 (901-1000)' or 'EG-1 (TOTAL)'
    raw = re.sub(r'\s*\([^)]*\)\s*$', '', raw).strip()
    # Digit-suffix marks with word suffix: LD-1-PLENUM → LD-1
    m = _COMPOUND_TAG_RE.match(raw)
    if m and m.group(1) and m.group(1) != raw:
        return m.group(1)
    # Letter-only marks with word suffix: F-PLENUM → F, F-SLOT DIFFUSER → F
    # (_COMPOUND_TAG_RE requires digits so can't handle these)
    m2 = re.match(r'^([A-Z]{1,4})-[A-Z]{2,}', raw)
    if m2:
        base = m2.group(1)
        if base.isalpha():   # no digits in base → safe to strip word suffix
            return base
    return raw


def _extract_cfm(s) -> int | None:
    """Extract first integer ≥10 from a string like 'CFM: 240' or '240'."""
    if s is None:
        return None
    m = re.search(r'\b(\d{2,6})\b', str(s))
    return int(m.group(1)) if m else None


# ── Team file parser ──────────────────────────────────────────────────────────

def _detect_data_columns(header_row: tuple) -> dict:
    """
    Map logical field names to 0-based column indices from the DATA sheet header.

    Handles the common naming variants seen across team Excel templates:
      TAG: TAG, MARK, UNIT TAG, SYMBOL
      NECK SIZE: NECK SIZE, NECK, SIZE
      MODULE SIZE: MODULE SIZE, MODULE, FACE SIZE
      DUCT SIZE: DUCT SIZE, DUCT
      CFM: CFM, AIR FLOW, AIRFLOW
      TYPE: TYPE, DESCRIPTION, SERVICE
      MOUNTING: MOUNTING, MOUNT
      BRAND: BRAND, MANUFACTURER, MAKE
      MODEL: MODEL, MODEL NUMBER
      PRODUCT: PRODUCT, ITEM, EQUIPMENT
    Falls back to the Busy Bees hard-coded positions if a field is not found.
    """
    _FIELD_ALIASES = {
        'tag':       ['TAG', 'MARK', 'UNIT TAG', 'SYMBOL', 'ID'],
        'neck':      ['NECK SIZE', 'NECK', 'SIZE'],
        'module':    ['MODULE SIZE', 'MODULE', 'FACE SIZE'],
        'duct':      ['DUCT SIZE', 'DUCT'],
        'cfm':       ['CFM', 'AIR FLOW', 'AIRFLOW', 'FLOW'],
        'type':      ['TYPE', 'DESCRIPTION', 'SERVICE'],
        'mount':     ['MOUNTING', 'MOUNT'],
        'brand':     ['BRAND', 'MANUFACTURER', 'MAKE'],
        'model':     ['MODEL', 'MODEL NUMBER', 'MODEL NO'],
        'product':   ['PRODUCT', 'ITEM', 'EQUIPMENT', 'PRODUCT TYPE'],
        'qty_direct':['QTY', 'COUNT', 'QUANTITY'],  # Haldeman pivot sheet
    }
    _FALLBACK = {
        'product': 0, 'tag': 2, 'brand': 4, 'model': 5,
        'neck': 6, 'module': 7, 'duct': 8, 'cfm': 9,
        'type': 10, 'mount': 11,
    }

    col_map: dict[str, int] = {}
    for ci, cell in enumerate(header_row or []):
        if not cell:
            continue
        cell_u = str(cell).strip().upper()
        for field, aliases in _FIELD_ALIASES.items():
            if field not in col_map:
                for alias in aliases:
                    if alias in cell_u:
                        col_map[field] = ci
                        break

    # Fill any undetected fields with fallback hard-coded positions
    for field, idx in _FALLBACK.items():
        col_map.setdefault(field, idx)

    return col_map


# Product-type keywords that identify GRD items in the PRODUCT column.
# Rows whose PRODUCT does not contain any of these (case-insensitive) are
# classified as equipment (FCU, PTAC, HP, etc.) and excluded from GRD
# comparison when grd_only=True.
_GRD_PRODUCT_KEYWORDS = (
    'GRD', 'GRILLE', 'REGISTER', 'DIFFUSER', 'LINEAR', 'LOUVER',
    'AIR DEVICE', 'AD-', 'SUPPLY', 'RETURN', 'EXHAUST',
)

# Product keywords that always indicate equipment (hard-exclude even if a
# GRD keyword is also present — e.g. "SUPPLY FAN" should not be GRD).
_EQUIPMENT_PRODUCT_KEYWORDS = (
    'FAN COIL', 'FCU', 'RTU', 'PTAC', 'HEAT PUMP', 'AHU',
    'CONDENSING', 'SPLIT SYSTEM', 'ERV', 'UNIT HEATER',
    'EXHAUST FAN',   # EF-1 type — fan, not a diffuser/grille
    'FIRE PROTECTION', 'SPRINKLER',
    'ELECTRIC UNIT', 'ELECTRIC HEATER',
)


def _is_grd_product(product_str: str) -> bool:
    """Return True if the PRODUCT column value identifies a GRD item."""
    if not product_str:
        return False
    pu = product_str.upper()
    # Hard-exclude equipment even if a GRD keyword appears
    for kw in _EQUIPMENT_PRODUCT_KEYWORDS:
        if kw in pu:
            return False
    for kw in _GRD_PRODUCT_KEYWORDS:
        if kw in pu:
            return True
    return False


def parse_team(path: str, grd_only: bool = True) -> list[dict]:
    """
    Parse team's completed takeoff Excel using the DATA sheet as ground truth.

    The DATA sheet has one raw row per physical instance (no formulas) so:
        qty = row-count per (tag_norm, neck_norm)

    Parameters
    ----------
    path : str
        Path to team Excel file.
    grd_only : bool
        When True (default), only include rows whose PRODUCT column identifies
        a GRD item (grille, register, diffuser, louver, linear diffuser).
        Rows with PRODUCT like FCU, PTAC, HP, RTU are excluded.
        Set to False to return all DATA rows regardless of product type.

    Column positions are detected from the header row to support different
    team Excel templates. Falls back to Busy Bees hard-coded positions when
    a column header is not recognized.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)

    # ── Sheet selection ───────────────────────────────────────────────────────
    # Primary: DATA sheet (one row per physical instance, qty += 1 per row)
    # Fallback: Triune Takeoff Haldeman / TAKEOFF  (aggregated pivot, use QTY col)
    _HALDEMAN_SHEETS = ('Triune Takeoff Haldeman', 'TAKEOFF', 'Takeoff',
                        'Triune Takeoff')
    _haldeman_mode = False

    if 'DATA' in wb.sheetnames:
        ws = wb['DATA']
    else:
        alt = next((s for s in _HALDEMAN_SHEETS if s in wb.sheetnames), None)
        if alt is None:
            raise ValueError(
                f"'DATA' sheet not found in {path}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[alt]
        _haldeman_mode = True   # QTY column contains actual unit count per row

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    cols = _detect_data_columns(rows[0])
    _qty_col = cols.get('qty_direct', None)   # used only in Haldeman mode

    agg: dict[tuple, dict] = {}   # (tag_norm, neck_norm) -> aggregated record

    def _get(row_: tuple, field: str) -> str:
        ci = cols.get(field, -1)
        return str(row_[ci] or '').strip() if 0 <= ci < len(row_) else ''

    for row in rows[1:]:   # skip header row
        if not row or all(c is None for c in row):
            continue

        tag = _norm_tag(_get(row, 'tag'))
        if not tag:
            continue

        product = _get(row, 'product')
        if grd_only and product and not _is_grd_product(product):
            continue   # exclude equipment rows from GRD comparison

        # Haldeman pivot sheets have "Total" rows (product ends with 'Tot' or
        # 'Total') that should not be counted as instances.
        if _haldeman_mode and product.upper().endswith(('TOT', 'TOTAL')):
            continue

        neck_raw   = _get(row, 'neck')
        module_raw = _get(row, 'module')
        duct_raw   = _get(row, 'duct')
        type_val   = _get(row, 'type')
        mount_val  = _get(row, 'mount')
        brand_val  = _get(row, 'brand')
        model_val  = _get(row, 'model')
        product    = _get(row, 'product')
        cfm_ci     = cols.get('cfm', -1)
        cfm_raw    = row[cfm_ci] if 0 <= cfm_ci < len(row) else None

        neck_n = _norm_dim(neck_raw)
        mod_n  = _norm_dim(module_raw)
        key    = (tag, neck_n)

        if key not in agg:
            agg[key] = {
                'tag':         tag,
                'neck_raw':    neck_raw,
                'neck_norm':   neck_n,
                'module_raw':  module_raw,
                'module_norm': mod_n,
                'duct_raw':    duct_raw,
                'qty':         0,
                'type':        type_val,
                'mounting':    mount_val,
                'brand':       brand_val,
                'model':       model_val,
                'product':     product or '',
                'cfm':         None,
                '_cfm_sum':    0,
                '_cfm_hits':   0,
            }

        qty_ci  = cols.get('qty_direct', -1)
        qty_val = row[qty_ci] if 0 <= qty_ci < len(row) else None
        try:
            qty_int = int(qty_val or 0)
        except (ValueError, TypeError):
            qty_int = 0

        if _haldeman_mode:
            # Haldeman pivot: QTY column holds the actual unit count
            agg[key]['qty'] += max(qty_int, 1)
        elif qty_int > 1:
            # DATA sheet with explicit Count/QTY column (e.g. UT Health style).
            # When count > 1 the row is aggregated, not a per-instance record.
            agg[key]['qty'] += qty_int
        else:
            agg[key]['qty'] += 1   # each DATA row is one physical instance

        if cfm_raw and str(cfm_raw) not in ('.', '', 'None'):
            try:
                agg[key]['_cfm_sum']  += int(cfm_raw)
                agg[key]['_cfm_hits'] += 1
            except (ValueError, TypeError):
                pass

    records = []
    for r in sorted(agg.values(), key=lambda x: (x['tag'], x['neck_norm'])):
        if r['_cfm_hits'] > 0:
            r['cfm'] = r['_cfm_sum']
        del r['_cfm_sum'], r['_cfm_hits']
        records.append(r)

    total_units = sum(r['qty'] for r in records)
    filter_note = " (GRD products only)" if grd_only else " (all products)"
    print(f"  Team: {len(records)} GRD rows, {total_units} units  "
          f"({Path(path).name} — DATA sheet{filter_note})")
    return records


# ── Our file parser ───────────────────────────────────────────────────────────

def parse_ours(path: str) -> list[dict]:
    """
    Parse our generated takeoff Excel.
    Finds the GRD section in 'Triune Takeoff' sheet.

    GRD section layout (written by write_excel in takeoff_cli.py):
      Section header: col A  = 'GRILLES / REGISTERS / DIFFUSERS (text-layer)'
      Sub-headers:    col A–K = standard, col L–N = pricing
      Data rows:      col A=mark(blank=carry), col B=brand, col C=model,
                      col D=qty, col E=tag, col F=neck_disp, col G=module,
                      col J=mounting, col K=remark ('CFM: N')
      End sentinel:   col A  = 'GRD TOTAL'
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)

    if 'Triune Takeoff' not in wb.sheetnames:
        raise ValueError(f"'Triune Takeoff' sheet not found in {path}")

    ws    = wb['Triune Takeoff']
    rows  = list(ws.iter_rows(values_only=True))

    # Locate GRD section header row (0-indexed)
    grd_header_idx = None
    for i, row in enumerate(rows):
        if row[0] and 'GRILLE' in str(row[0]).upper():
            grd_header_idx = i
            break

    if grd_header_idx is None:
        print(f"  Ours: GRD section not found in {Path(path).name}")
        return []

    # Data starts two rows after section header (skip section header + sub-headers)
    records      = []
    current_mark = None

    for row in rows[grd_header_idx + 2:]:
        if all(c is None for c in row):
            continue
        col_a = str(row[0] or '').strip()
        if 'GRD TOTAL' in col_a.upper():
            break

        # col A = mark (blank when same as previous — forward-fill)
        if col_a:
            current_mark = col_a
        mark = current_mark
        if not mark:
            continue

        qty_val    = row[3]    # col D
        tag_val    = row[4]    # col E  (always explicit)
        neck_raw   = str(row[5] or '').strip()   # col F
        module_raw = str(row[6] or '').strip()   # col G
        mount_val  = str(row[9] or '').strip()   # col J
        remark_val = row[10]                     # col K
        brand_val  = str(row[1] or '').strip()   # col B
        model_val  = str(row[2] or '').strip()   # col C

        if not qty_val:
            continue

        tag_n  = _norm_tag(tag_val or mark)
        neck_n = _norm_dim(neck_raw)
        mod_n  = _norm_dim(module_raw)
        cfm    = _extract_cfm(remark_val)

        records.append({
            'tag':         tag_n,
            'neck_raw':    neck_raw,
            'neck_norm':   neck_n,
            'module_raw':  module_raw,
            'module_norm': mod_n,
            'duct_raw':    '',
            'qty':         int(qty_val),
            'type':        '',
            'mounting':    mount_val,
            'brand':       brand_val,
            'model':       model_val,
            'cfm':         cfm,
        })

    print(f"  Ours: {len(records)} GRD rows from {Path(path).name}")
    return records


# ── Comparison engine ─────────────────────────────────────────────────────────

def compare(team: list[dict], ours: list[dict]) -> dict:
    """
    Compare two GRD record lists keyed on (tag_norm, neck_norm).
    Returns dict with matched / missing / extra / diffs / summary.
    """
    team_by_key: dict[tuple, dict] = {}
    for r in team:
        key = (r['tag'], r['neck_norm'])
        team_by_key[key] = r

    ours_by_key: dict[tuple, dict] = {}
    for r in ours:
        key = (r['tag'], r['neck_norm'])
        ours_by_key[key] = r

    all_keys = sorted(
        set(team_by_key) | set(ours_by_key),
        key=lambda k: (k[0], k[1]),
    )

    matched: list[dict] = []
    missing: list[dict] = []   # in team, not in ours
    extra:   list[dict] = []   # in ours, not in team
    diffs:   list[dict] = []   # matched but field-level differences

    for key in all_keys:
        in_t = key in team_by_key
        in_o = key in ours_by_key

        if in_t and not in_o:
            missing.append({**team_by_key[key], 'status': 'MISSING'})

        elif in_o and not in_t:
            extra.append({**ours_by_key[key], 'status': 'EXTRA'})

        else:
            t = team_by_key[key]
            o = ours_by_key[key]
            field_diffs: list[str] = []

            # QTY
            if t['qty'] != o['qty']:
                field_diffs.append(
                    f"QTY: team={t['qty']} ours={o['qty']} "
                    f"(Δ={o['qty']-t['qty']:+d})"
                )
            # Module size
            if t['module_norm'] and o['module_norm']:
                if t['module_norm'] != o['module_norm']:
                    field_diffs.append(
                        f"MODULE: team={t['module_raw']} ours={o['module_raw']}"
                    )
            # CFM — only flag when both sides have a value
            if t['cfm'] and o['cfm']:
                delta = abs(t['cfm'] - o['cfm'])
                if delta > 0:
                    field_diffs.append(
                        f"CFM: team={t['cfm']} ours={o['cfm']} "
                        f"(Δ={o['cfm']-t['cfm']:+d})"
                    )

            merged = {
                **t,
                'our_qty':         o['qty'],
                'our_neck_raw':    o['neck_raw'],
                'our_module_raw':  o['module_raw'],
                'our_module_norm': o['module_norm'],
                'our_cfm':         o['cfm'],
                'our_brand':       o['brand'],
                'our_model':       o['model'],
                'our_mounting':    o['mounting'],
                'field_diffs':     field_diffs,
                'status':          'DIFF' if field_diffs else 'MATCH',
            }
            matched.append(merged)
            if field_diffs:
                diffs.append(merged)

    # --- Tag-level aggregation ---
    tag_team: dict[str, int] = defaultdict(int)
    for r in team:
        tag_team[r['tag']] += r['qty']

    tag_ours: dict[str, int] = defaultdict(int)
    for r in ours:
        tag_ours[r['tag']] += r['qty']

    all_tags = sorted(set(tag_team) | set(tag_ours))
    tag_rows = []
    for tag in all_tags:
        t_qty = tag_team.get(tag, 0)
        o_qty = tag_ours.get(tag, 0)
        tag_rows.append({
            'tag':     tag,
            'team':    t_qty,
            'ours':    o_qty,
            'delta':   o_qty - t_qty,
            'status':  ('MATCH' if t_qty == o_qty
                        else 'EXTRA' if t_qty == 0
                        else 'MISSING' if o_qty == 0
                        else 'DIFF'),
        })

    # --- Summary stats ---
    team_total = sum(r['qty'] for r in team)
    ours_total = sum(r['qty'] for r in ours)

    # Recall = matched quantity / team total
    matched_qty = sum(
        min(team_by_key[k]['qty'], ours_by_key[k]['qty'])
        for k in set(team_by_key) & set(ours_by_key)
    )
    recall    = round(matched_qty / max(team_total, 1) * 100, 1)
    precision = round(matched_qty / max(ours_total, 1) * 100, 1)

    summary = {
        'team_rows':    len(team),
        'ours_rows':    len(ours),
        'team_total':   team_total,
        'ours_total':   ours_total,
        'matched_rows': sum(1 for r in matched if r['status'] == 'MATCH'),
        'diff_rows':    len(diffs),
        'missing_rows': len(missing),
        'extra_rows':   len(extra),
        'qty_delta':    ours_total - team_total,
        'recall_pct':   recall,
        'precision_pct': precision,
    }

    return {
        'matched': matched,
        'missing': missing,
        'extra':   extra,
        'diffs':   diffs,
        'tag_rows': tag_rows,
        'summary': summary,
    }


# ── Excel report writer ───────────────────────────────────────────────────────

def _hdr(ws, row: int, col: int, value, wide: bool = False) -> None:
    """Write a styled header cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font   = _BOLD
    c.fill   = _FILL['header']
    c.border = _THIN
    if wide:
        c.alignment = Alignment(wrap_text=True, vertical='top')


def _dat(ws, row: int, col: int, value, fill_key: str | None = None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.border = _THIN
    if fill_key:
        c.fill = _FILL[fill_key]


def _set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_summary(wb: openpyxl.Workbook, result: dict) -> None:
    ws = wb.create_sheet('Summary')
    s  = result['summary']

    # Title
    ws.merge_cells('A1:D1')
    tc = ws['A1']
    tc.value     = 'GRD TAKEOFF COMPARISON REPORT'
    tc.font      = _WHITE
    tc.fill      = _FILL['title']
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    rows = [
        ('METRIC',                 'TEAM',             'OURS',            'DELTA / NOTE'),
        ('BOM rows (unique lines)', s['team_rows'],     s['ours_rows'],    s['ours_rows']  - s['team_rows']),
        ('Total units (qty sum)',   s['team_total'],    s['ours_total'],   s['qty_delta']),
        ('Rows: exact match',      s['matched_rows'],  '',                ''),
        ('Rows: field diff',       s['diff_rows'],     '',                ''),
        ('Rows: missing from ours',s['missing_rows'],  '',                ''),
        ('Rows: extra in ours',    s['extra_rows'],    '',                ''),
        ('Recall  (%)',            '',                 f"{s['recall_pct']}%",    'matched_qty / team_total'),
        ('Precision (%)',          '',                 f"{s['precision_pct']}%", 'matched_qty / ours_total'),
    ]

    for ri, row_data in enumerate(rows, 2):
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _THIN
            if ri == 2:
                c.font = _BOLD
                c.fill = _FILL['header']
            # Colour-code delta column
            if ci == 4 and isinstance(val, int):
                if val > 0:   c.fill = _FILL['extra']
                elif val < 0: c.fill = _FILL['missing']
                else:         c.fill = _FILL['match']

    # Status legend
    legend_start = len(rows) + 3
    ws.cell(row=legend_start, column=1, value='STATUS LEGEND').font = _BOLD
    for offset, (key, label) in enumerate([
        ('match',   'MATCH   — exact qty, neck, module'),
        ('diff',    'DIFF    — matched key, field value differs'),
        ('missing', 'MISSING — in team takeoff but absent from ours'),
        ('extra',   'EXTRA   — in our output but not in team takeoff'),
    ], 1):
        c = ws.cell(row=legend_start + offset, column=1, value=label)
        c.fill   = _FILL[key]
        c.border = _THIN

    _set_widths(ws, [32, 18, 18, 38])


def _write_grd_comparison(wb: openpyxl.Workbook, result: dict) -> None:
    ws = wb.create_sheet('GRD_Comparison')

    headers = [
        'STATUS', 'TAG', 'NECK (NORM)',
        'TEAM QTY', 'OURS QTY', 'QTY Δ',
        'TEAM NECK', 'OURS NECK',
        'TEAM MODULE', 'OURS MODULE',
        'TEAM CFM', 'OURS CFM',
        'TEAM TYPE', 'DIFFERENCES',
    ]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    row = 2
    all_rows = (
        result['matched'] +
        result['missing'] +
        result['extra']
    )
    # Sort by (tag, neck_norm) for readability
    all_rows.sort(key=lambda r: (r['tag'], r['neck_norm']))

    for r in all_rows:
        status   = r['status']
        fill_key = {'MATCH': 'match', 'DIFF': 'diff',
                    'MISSING': 'missing', 'EXTRA': 'extra'}.get(status)

        our_qty     = r.get('our_qty',     r['qty']   if status == 'EXTRA' else '')
        our_neck    = r.get('our_neck_raw', r['neck_raw'] if status == 'EXTRA' else '')
        our_module  = r.get('our_module_raw', r['module_raw'] if status == 'EXTRA' else '')
        our_cfm     = r.get('our_cfm',     r.get('cfm') if status == 'EXTRA' else '')

        team_qty    = r['qty']   if status != 'EXTRA' else ''
        team_neck   = r['neck_raw']   if status != 'EXTRA' else ''
        team_module = r['module_raw'] if status != 'EXTRA' else ''
        team_cfm    = r.get('cfm')    if status != 'EXTRA' else ''

        qty_delta = ''
        if isinstance(our_qty, int) and isinstance(team_qty, int):
            qty_delta = our_qty - team_qty

        diffs_str = '; '.join(r.get('field_diffs', []))

        vals = [
            status, r['tag'], r['neck_norm'],
            team_qty, our_qty, qty_delta,
            team_neck, our_neck,
            team_module, our_module,
            team_cfm, our_cfm,
            r.get('type', ''), diffs_str,
        ]
        for ci, v in enumerate(vals, 1):
            _dat(ws, row, ci, v, fill_key)
            # Extra colouring for delta column
            if ci == 6 and isinstance(v, int):
                if v > 0:   ws.cell(row=row, column=ci).fill = _FILL['extra']
                elif v < 0: ws.cell(row=row, column=ci).fill = _FILL['missing']

        row += 1

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
    ws.freeze_panes    = 'A2'
    _set_widths(ws, [10, 6, 12, 10, 10, 8, 12, 12, 14, 14, 10, 10, 22, 50])


def _write_tag_counts(wb: openpyxl.Workbook, result: dict) -> None:
    ws = wb.create_sheet('Tag_Counts')

    headers = ['TAG', 'TEAM TOTAL QTY', 'OURS TOTAL QTY', 'DELTA', 'STATUS']
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    # Totals row at bottom
    for ri, r in enumerate(result['tag_rows'], 2):
        fill_key = {'MATCH': 'match', 'DIFF': 'diff',
                    'MISSING': 'missing', 'EXTRA': 'extra'}.get(r['status'])
        vals = [r['tag'], r['team'], r['ours'], r['delta'], r['status']]
        for ci, v in enumerate(vals, 1):
            _dat(ws, ri, ci, v, fill_key)

    # Summary totals
    last = len(result['tag_rows']) + 2
    ws.cell(row=last, column=1, value='TOTAL').font = _BOLD
    ws.cell(row=last, column=2,
            value=result['summary']['team_total']).font = _BOLD
    ws.cell(row=last, column=3,
            value=result['summary']['ours_total']).font = _BOLD
    ws.cell(row=last, column=4,
            value=result['summary']['qty_delta']).font = _BOLD
    for ci in range(1, 6):
        ws.cell(row=last, column=ci).border = _THIN
        ws.cell(row=last, column=ci).fill   = _FILL['header']

    _set_widths(ws, [8, 16, 16, 10, 10])


def _write_detail_sheet(wb: openpyxl.Workbook, name: str,
                        records: list[dict], fill_key: str) -> None:
    if not records:
        ws = wb.create_sheet(name)
        ws['A1'].value = f'No {name.lower()} rows.'
        return

    ws      = wb.create_sheet(name)
    headers = ['TAG', 'NECK (NORM)', 'NECK RAW', 'QTY',
               'MODULE RAW', 'TYPE', 'MOUNTING', 'CFM', 'BRAND', 'MODEL']
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    for ri, r in enumerate(records, 2):
        vals = [
            r['tag'], r['neck_norm'], r['neck_raw'], r['qty'],
            r['module_raw'], r.get('type', ''), r.get('mounting', ''),
            r.get('cfm', ''), r.get('brand', ''), r.get('model', ''),
        ]
        for ci, v in enumerate(vals, 1):
            _dat(ws, ri, ci, v, fill_key)

    _set_widths(ws, [8, 12, 12, 6, 14, 22, 12, 8, 14, 10])


def _write_field_diffs(wb: openpyxl.Workbook, result: dict) -> None:
    diffs = result['diffs']
    ws = wb.create_sheet('Field_Diffs')

    if not diffs:
        ws['A1'].value = 'No field-level differences found.'
        return

    headers = ['TAG', 'NECK (NORM)',
               'TEAM QTY', 'OURS QTY',
               'TEAM MODULE', 'OURS MODULE',
               'TEAM CFM', 'OURS CFM',
               'DIFFERENCES']
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    for ri, r in enumerate(diffs, 2):
        vals = [
            r['tag'], r['neck_norm'],
            r['qty'], r['our_qty'],
            r['module_raw'], r['our_module_raw'],
            r.get('cfm', ''), r.get('our_cfm', ''),
            '; '.join(r['field_diffs']),
        ]
        for ci, v in enumerate(vals, 1):
            _dat(ws, ri, ci, v, 'diff')

    _set_widths(ws, [8, 12, 10, 10, 14, 14, 10, 10, 60])


def write_report(result: dict, out_path: str) -> None:
    wb = openpyxl.Workbook()
    del wb['Sheet']   # remove default sheet

    _write_summary(wb, result)
    _write_grd_comparison(wb, result)
    _write_tag_counts(wb, result)
    _write_detail_sheet(wb, 'Missing_Rows', result['missing'], 'missing')
    _write_detail_sheet(wb, 'Extra_Rows',   result['extra'],   'extra')
    _write_field_diffs(wb, result)

    wb.save(str(out_path))
    print(f"\n  Report saved: {out_path}")


# ── CLI entry point ───────────────────────────────────────────────────────────

_DEFAULT_TEAM = (
    r'C:\Users\91739\Downloads\@09-29 Busy Bees Queen Creek'
    r'\@09-29 Busy Bees Queen Creek\Completed Takeoff'
    r'\Takeoff_Busy Bees Queen Creek.xlsx'
)
_DEFAULT_OURS = (
    r'C:\Users\91739\Desktop\hvac-takeoff-tool\output\busy_bees_test'
    r'\Nikki_Schroeder_09-29 Busy Bees Queen Creek_takeoff.xlsx'
)
_DEFAULT_OUT = (
    r'C:\Users\91739\Desktop\hvac-takeoff-tool\output\busy_bees_test'
    r'\comparison_report.xlsx'
)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--team', default=_DEFAULT_TEAM,
                    help='Path to team completed takeoff Excel')
    ap.add_argument('--ours', default=_DEFAULT_OURS,
                    help='Path to our generated takeoff Excel')
    ap.add_argument('--out',  default=_DEFAULT_OUT,
                    help='Output path for comparison_report.xlsx')
    args = ap.parse_args()

    print(f"Parsing team file:  {Path(args.team).name}")
    team = parse_team(args.team)

    print(f"Parsing ours file:  {Path(args.ours).name}")
    ours = parse_ours(args.ours)

    print("\nRunning comparison...")
    result = compare(team, ours)
    s = result['summary']

    SEP = '-' * 52
    print(f"\n{SEP}")
    print(f"  {'Metric':<28}  {'Team':>6}  {'Ours':>6}  {'Delta':>6}")
    print(SEP)
    print(f"  {'BOM rows':<28}  {s['team_rows']:>6}  {s['ours_rows']:>6}  "
          f"{s['ours_rows']-s['team_rows']:>+6}")
    print(f"  {'Total units (qty)':<28}  {s['team_total']:>6}  "
          f"{s['ours_total']:>6}  {s['qty_delta']:>+6}")
    print(SEP)
    print(f"  Exact matches:       {s['matched_rows']}")
    print(f"  Field diffs:         {s['diff_rows']}")
    print(f"  Missing (team only): {s['missing_rows']}")
    print(f"  Extra   (ours only): {s['extra_rows']}")
    print(SEP)
    print(f"  Recall:    {s['recall_pct']:>5.1f}%")
    print(f"  Precision: {s['precision_pct']:>5.1f}%")
    print(SEP)

    print("\nTag-level summary:")
    print(f"  {'TAG':<6} {'TEAM':>5} {'OURS':>5} {'DELTA':>6}  STATUS")
    for r in result['tag_rows']:
        print(f"  {r['tag']:<6} {r['team']:>5} {r['ours']:>5} "
              f"{r['delta']:>+6}  {r['status']}")

    print("\nWriting report...")
    write_report(result, args.out)


if __name__ == '__main__':
    main()
